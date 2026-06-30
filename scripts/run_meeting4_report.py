"""
Meeting 4 Progress Report — Algorithm Comparison and HET Application.

Purpose
-------
This script generates the visual and tabular outputs for the meeting
progress report. It demonstrates the three quantum linear system
algorithms (HHL, VQLS, QSVT) implemented in this repository, applied to
the 1-D Poisson equation and the Hall Effect Thruster plasma modelling
problem.

Report structure
----------------
Section 1 — Algorithm Comparison on the 1-D Poisson Equation.
    All three solvers are run on the same N=4 and N=8 test cases with
    the fS source function and homogeneous boundary conditions, where
    the analytical solution is available for exact error quantification.
    Output: Figure 1 (solution profiles and errors), Table 1 (metrics).

Section 2 — QSVT Circuit Complexity Analysis.
    The QSVT polynomial degree, circuit depth, and qubit count are
    computed analytically for N in {4, 8, 16, 32} and compared against
    HHL circuit depth estimates. This demonstrates the O(kappa) vs
    O(kappa^2) scaling advantage of QSVT.
    Output: Figure 2 (scaling plots), Table 2 (complexity metrics).

Section 3 — HET Plasma Application (1-D).
    HHL and VQLS are applied to the physical HET Poisson problem with
    the Boeuf-Garrigues (1998) parameters. The electric field profile
    is compared qualitatively against the published result.
    Output: Figure 3 (potential and electric field profiles), Table 3.

Section 4 — HET Plasma Application (2-D).
    Thomas and VQLS are applied to the 2-D HET sinusoidal problem with
    the analytical solution. Contour plots of the potential and error
    fields are generated.
    Output: Figure 4 (contour plots), Table 4.

Section 5 — 2-D QSVT 
   Explain ... # TODO: Add description of Section 5

Output artefacts
----------------
All figures are saved as PDF and PNG to results/meeting_report/.
All tables are saved as a single Excel workbook with one sheet per table.
A plain-text summary is printed to the console.

Runtime estimate
----------------
Section 1 (QSVT N=4):          ~10-15 minutes
Section 1 (HHL/VQLS N=4,8):    ~5-10 minutes
Section 2 (analytical only):    < 1 minute
Section 3 (HET 1-D, N=8):      ~10-15 minutes
Section 4 (HET 2-D, N=4):      ~10-15 minutes
Total:                          ~35-55 minutes

References
----------
Ghafourpour & Laizet, Phys. Rev. Applied 24, 024032 (2025).
Bravo-Prieto et al., Quantum 7, 1188 (2023).
Harrow, Hassidim & Lloyd, Phys. Rev. Lett. 103, 150502 (2009).
Gilyen et al., STOC 2019, pp. 193-204.
Boeuf & Garrigues, J. Appl. Phys. 84(7), 3541-3554 (1998).
"""
from __future__ import annotations

import time
import warnings
import sys
from pathlib import Path

# ── System Path Resolution ────────────────────────────────────────────────────

# Dynamically resolve the project root directory (one level up from this script)
# and append it to the system path to enable absolute imports.
project_root = Path(__file__).resolve().parents[1]
if str(project_root) not in sys.path:
    sys.path.append(str(project_root))

import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt
import numpy as np

# -- Excel export (requires openpyxl) --------------------------------
try:
    import openpyxl
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False
    warnings.warn(
        "openpyxl not installed; Excel export will be skipped. "
        "Install via: pip install openpyxl",
        ImportWarning,
    )

from core.config import SimConfig1D, SimConfig2D
from core.exact_solutions import EXACT_SOLUTIONS, HET_EXACT_SOLUTIONS
from core.het_config import HETConfig
from problems.het_plasma_1d import HETPoissonProblem1D
from problems.het_plasma_2d import HETConfig2D, HETSinusoidalProblem2D
from problems.poisson_1d import PoissonProblem1D
from solvers.classical.thomas import thomas_solve, thomas_solve_system
from solvers.classical.thomas_2d import thomas_solve_2d
from solvers.quantum.hhl_1d import hhl_solve, hhl_solve_system
from solvers.quantum.block_encoding import subnormalisation_factor
from solvers.quantum.qsp_angles import polynomial_degree_estimate
from solvers.quantum.qsvt_1d import QSVTConfig, qsvt_solve
from solvers.quantum.result import QSVTSolverResult
from solvers.quantum.vqls_1d import VQLSConfig1D, vqls_solve, vqls_solve_system
from solvers.quantum.vqls_2d import VQLSConfig2D, vqls_solve_2d
from solvers.quantum.qsvt_2d import QSVTConfig2D, qsvt_solve_2d

RESULTS_DIR = Path("results/meeting_report")
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

# -- Matplotlib global style --------------------------------------------------
plt.rcParams.update({
    "font.family":    "serif",
    "font.size":      11,
    "axes.labelsize": 12,
    "axes.titlesize": 11,
    "legend.fontsize": 9,
    "figure.dpi":     140,
    "lines.linewidth": 1.8,
    "lines.markersize": 6,
})

COLOURS = {
    "thomas":     "#2ca02c",
    "hhl":        "#1f77b4",
    "vqls":       "#d62728",
    "qsvt":       "#9467bd",
    "analytical": "#000000",
}
MARKERS = {
    "thomas": "o",
    "hhl":    "s",
    "vqls":   "^",
    "qsvt":   "D",
}


# -- Utility functions --------------------------------------------------------

def _rel_err_pct(u: np.ndarray, ref: np.ndarray) -> np.ndarray:
    """
    Pointwise absolute relative error in percent.

    Nodes where |ref| < 1e-4 * max|ref| are masked to NaN to prevent
    division by near-zero values from inflating the metric.

    Parameters
    ----------
    u : np.ndarray, shape (N,)
    ref : np.ndarray, shape (N,)

    Returns
    -------
    err : np.ndarray, shape (N,)
    """
    scale = np.max(np.abs(ref))
    mask  = np.abs(ref) > 1e-4 * scale
    return np.where(mask, np.abs(u - ref) / np.abs(ref) * 100.0, np.nan)


def _max_rel_err(u: np.ndarray, ref: np.ndarray) -> float:
    """Maximum relative error in percent, excluding masked nodes."""
    err   = _rel_err_pct(u, ref)
    valid = err[~np.isnan(err)]
    return float(np.max(valid)) if valid.size > 0 else float("nan")


def _electric_field_1d(
    phi_int  : np.ndarray,
    alpha_bc : float,
    phi_0    : float,
    L        : float,
    N        : int,
) -> tuple[np.ndarray, np.ndarray]:
    """
    Recover the physical electric field E(x) [V/m] from the
    non-dimensional interior potential via second-order centred
    finite differences.

    Parameters
    ----------
    phi_int : np.ndarray, shape (N,)
    alpha_bc : float
    phi_0 : float
    L : float
    N : int

    Returns
    -------
    x_full : np.ndarray, shape (N+2,)
    E_phys : np.ndarray, shape (N+2,)
    """
    dx           = 1.0 / (N + 1)
    phi_full     = np.zeros(N + 2)
    phi_full[0]  = alpha_bc
    phi_full[1:N+1] = phi_int
    phi_full[N+1]   = 0.0
    x_full       = np.linspace(0.0, 1.0, N + 2)
    E_nd         = np.zeros(N + 2)
    E_nd[1:-1]   = -(phi_full[2:] - phi_full[:-2]) / (2.0 * dx)
    E_nd[0]      = -(phi_full[1]  - phi_full[0])   / dx
    E_nd[-1]     = -(phi_full[-1] - phi_full[-2])  / dx
    return x_full, E_nd * phi_0 / L


def _section_header(title: str, index: int) -> None:
    """Print a formatted section header to the console."""
    print(f"\n{'═'*68}")
    print(f"  SECTION {index} — {title}")
    print(f"{'═'*68}")


def _solver_row(
    label    : str,
    rel_err  : float,
    abs_err  : float,
    residual : float,
    elapsed  : float,
    extra    : str = "",
) -> None:
    """Print a single formatted solver result row."""
    print(
        f"  {label:<10} {rel_err:>10.3f}%  {abs_err:>12.4e}  "
        f"{residual:>12.4e}  {elapsed:>8.2f}s  {extra}"
    )


# ============================================================================
# Section 1 — Algorithm comparison on the 1-D Poisson equation
# ============================================================================

def run_section_1() -> dict:
    """
    Run all three quantum solvers (HHL, VQLS, QSVT) on the 1-D Poisson
    equation with the fS source function and homogeneous BCs.

    The analytical solution phi(x) = -sin(pi*x)/pi^2 provides an exact
    reference for quantitative error assessment at both N=4 and N=8.
    QSVT is run at N=4 only due to circuit depth constraints; HHL and
    VQLS are run at both N=4 and N=8.

    Returns
    -------
    dict
        Nested dictionary keyed by N, then by solver label, containing
        solution vectors, timing, and error metrics.
    """
    _section_header("Algorithm Comparison — 1-D Poisson, fS Source", 1)
    print(f"  {'Solver':<10} {'MaxRelErr':>10}  {'MaxAbsErr':>12}  "
          f"{'Residual':>12}  {'Time':>8}")
    print(f"  {'─'*60}")

    # VQLS configuration: 6 layers, 3 restarts.
    vqls_cfg = VQLSConfig1D(
        n_layers    = 6,
        optimiser   = "COBYLA",
        max_iter    = 300,
        tol         = 1e-6,
        random_seed = 42,
        verbose     = False,
    )

    # QSVT configuration: epsilon=0.05 for tractable circuit depth at N=4.
    qsvt_cfg = QSVTConfig(
        epsilon      = 0.05,
        angle_method = "auto",
        verbose      = False,
        max_degree   = 100,
    )

    results = {}

    for N in (4, 8):
        cfg     = SimConfig1D(N=N, epsilon=0.01, source_fn="fS")
        problem = PoissonProblem1D(cfg)
        u_exact = EXACT_SOLUTIONS["fS"](problem.x)

        print(f"\n  N={N}  (kappa={problem.kappa:.2f}):")
        results[N] = {"x": problem.x, "u_exact": u_exact, "cfg": cfg}

        # Thomas.
        t0       = time.perf_counter()
        r_thomas = thomas_solve(problem)
        t_thomas = time.perf_counter() - t0
        results[N]["thomas"] = {
            "u": r_thomas.u, "t": t_thomas,
            "rel": _max_rel_err(r_thomas.u, u_exact),
            "res": r_thomas.euclidean_residual,
        }
        _solver_row(
            "Thomas",
            _max_rel_err(r_thomas.u, u_exact),
            float(np.max(np.abs(r_thomas.u - u_exact))),
            r_thomas.euclidean_residual, t_thomas,
        )

        # HHL.
        t0     = time.perf_counter()
        r_hhl  = hhl_solve(problem)
        t_hhl  = time.perf_counter() - t0
        results[N]["hhl"] = {
            "u": r_hhl.u, "t": t_hhl,
            "rel": _max_rel_err(r_hhl.u, u_exact),
            "res": r_hhl.euclidean_residual,
        }
        _solver_row(
            "HHL",
            _max_rel_err(r_hhl.u, u_exact),
            float(np.max(np.abs(r_hhl.u - u_exact))),
            r_hhl.euclidean_residual, t_hhl,
        )

        # VQLS.
        t0      = time.perf_counter()
        r_vqls  = vqls_solve(problem, config=vqls_cfg)
        t_vqls  = time.perf_counter() - t0
        results[N]["vqls"] = {
            "u": r_vqls.u, "t": t_vqls,
            "rel": _max_rel_err(r_vqls.u, u_exact),
            "res": r_vqls.euclidean_residual,
            "cost": r_vqls.final_cost,
        }
        _solver_row(
            "VQLS",
            _max_rel_err(r_vqls.u, u_exact),
            float(np.max(np.abs(r_vqls.u - u_exact))),
            r_vqls.euclidean_residual, t_vqls,
            f"cost={r_vqls.final_cost:.2e}",
        )

        # QSVT — N=4 only (circuit depth constraint).
        if N == 4:
            t0      = time.perf_counter()
            r_qsvt  = qsvt_solve(problem, config=qsvt_cfg)
            t_qsvt  = time.perf_counter() - t0
            results[N]["qsvt"] = {
                "u":      r_qsvt.u,
                "t":      t_qsvt,
                "rel":    _max_rel_err(r_qsvt.u, u_exact),
                "res":    r_qsvt.euclidean_residual,
                "degree": r_qsvt.polynomial_degree,
                "depth":  r_qsvt.circuit_depth,
                "qubits": r_qsvt.n_qubits,
                "alpha":  r_qsvt.alpha,
                "kappa_eff": r_qsvt.kappa_effective,
            }
            _solver_row(
                "QSVT",
                _max_rel_err(r_qsvt.u, u_exact),
                float(np.max(np.abs(r_qsvt.u - u_exact))),
                r_qsvt.euclidean_residual, t_qsvt,
                f"deg={r_qsvt.polynomial_degree}, "
                f"depth={r_qsvt.circuit_depth}",
            )
        else:
            results[N]["qsvt"] = None
            print(
                f"  {'QSVT':<10} {'N/A (N=8 circuit depth exceeds':>10} "
                f"{'laptop threshold)':>12}"
            )

    return results


def plot_section_1(data: dict, save: bool = True) -> None:
    """
    Generate Figure 1: algorithm comparison on the 1-D Poisson equation.

    Layout (2 rows x 3 columns):
        Row 1 (N=4): solution profiles | relative error | algorithm labels
        Row 2 (N=8): solution profiles | relative error | timing bar chart

    Parameters
    ----------
    data : dict
        Output of run_section_1().
    save : bool
        If True, save to RESULTS_DIR.
    """
    fig = plt.figure(figsize=(15, 9))
    fig.suptitle(
        "Section 1 — Algorithm Comparison: HHL vs VQLS vs QSVT\n"
        r"1-D Poisson Equation, $f_S(x) = \sin(\pi x)$, "
        r"$u(0)=u(1)=0$, Analytical: $\tilde{\phi}=-\sin(\pi x)/\pi^2$",
        fontsize=12,
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.42, wspace=0.32)

    for row_idx, N in enumerate((4, 8)):
        d       = data[N]
        x       = d["x"]
        u_exact = d["u_exact"]

        x_full     = np.concatenate([[0.0], x, [1.0]])
        phi_exact  = np.concatenate([[0.0], u_exact,        [0.0]])
        phi_thomas = np.concatenate([[0.0], d["thomas"]["u"], [0.0]])
        phi_hhl    = np.concatenate([[0.0], d["hhl"]["u"],    [0.0]])
        phi_vqls   = np.concatenate([[0.0], d["vqls"]["u"],   [0.0]])

        # -- Panel 1: solution profiles ---------------------------------------
        ax = fig.add_subplot(gs[row_idx, 0])
        ax.plot(x_full, phi_exact,  color=COLOURS["analytical"],
                lw=2.5, label="Analytical", zorder=6)
        ax.plot(x_full, phi_thomas, color=COLOURS["thomas"],
                ls="--", marker=MARKERS["thomas"], ms=5, label="Thomas")
        ax.plot(x_full, phi_hhl,    color=COLOURS["hhl"],
                ls="-.", marker=MARKERS["hhl"],    ms=5, label="HHL")
        ax.plot(x_full, phi_vqls,   color=COLOURS["vqls"],
                ls=":",  marker=MARKERS["vqls"],   ms=5, label="VQLS")
        if d["qsvt"] is not None:
            phi_qsvt = np.concatenate([[0.0], d["qsvt"]["u"], [0.0]])
            ax.plot(x_full, phi_qsvt, color=COLOURS["qsvt"],
                    ls=(0, (3, 1, 1, 1)), marker=MARKERS["qsvt"],
                    ms=5, label="QSVT")
        ax.set_xlabel(r"$x$")
        ax.set_ylabel(r"$u(x)$")
        ax.set_title(f"Solution profiles (N={N})")
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

        # -- Panel 2: relative error (log scale) ------------------------------
        ax = fig.add_subplot(gs[row_idx, 1])
        for key, label in [
            ("thomas", "Thomas"),
            ("hhl",    "HHL"),
            ("vqls",   "VQLS"),
        ]:
            err = _rel_err_pct(d[key]["u"], u_exact)
            ax.semilogy(
                x, err,
                color=COLOURS[key],
                ls="--" if key == "thomas" else "-",
                marker=MARKERS[key], ms=5, label=label,
            )
        if d["qsvt"] is not None:
            err_q = _rel_err_pct(d["qsvt"]["u"], u_exact)
            ax.semilogy(
                x, err_q,
                color=COLOURS["qsvt"],
                ls=(0, (3, 1, 1, 1)),
                marker=MARKERS["qsvt"], ms=5, label="QSVT",
            )
        ax.set_xlabel(r"$x$")
        ax.set_ylabel("Relative error (%)")
        ax.set_title(f"Error vs analytical (N={N})")
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3, which="both")

        # -- Panel 3: timing bar chart (row 1) or summary (row 2) ------------
        ax = fig.add_subplot(gs[row_idx, 2])
        if row_idx == 0:
            # Bar chart of wall-clock times.
            solvers = ["Thomas", "HHL", "VQLS", "QSVT"]
            times   = [
                d["thomas"]["t"],
                d["hhl"]["t"],
                d["vqls"]["t"],
                d["qsvt"]["t"] if d["qsvt"] else 0.0,
            ]
            colours = [
                COLOURS["thomas"], COLOURS["hhl"],
                COLOURS["vqls"],   COLOURS["qsvt"],
            ]
            bars = ax.bar(solvers, times, color=colours, alpha=0.8,
                          edgecolor="black", linewidth=0.8)
            for bar, t in zip(bars, times):
                if t > 0:
                    ax.text(
                        bar.get_x() + bar.get_width() / 2,
                        bar.get_height() * 1.02,
                        f"{t:.1f}s", ha="center", va="bottom", fontsize=9,
                    )
            ax.set_ylabel("Wall-clock time [s]")
            ax.set_title(f"Computation time (N={N})")
            ax.set_yscale("log")
            ax.grid(True, alpha=0.3, axis="y")
        else:
            # Summary metrics table for N=8.
            rows = []
            for key, label in [
                ("thomas", "Thomas"), ("hhl", "HHL"), ("vqls", "VQLS")
            ]:
                rows.append([
                    label,
                    f"{d[key]['rel']:.3f}%",
                    f"{d[key]['res']:.2e}",
                    f"{d[key]['t']:.1f}s",
                ])
            rows.append(["QSVT", "N/A (N=8)", "—", "—"])
            tbl = ax.table(
                cellText  = rows,
                colLabels = ["Solver", "Max Rel. Err.", "Residual", "Time"],
                cellLoc   = "center",
                loc       = "center",
            )
            tbl.auto_set_font_size(False)
            tbl.set_fontsize(9)
            tbl.scale(1.0, 1.6)
            for col in range(4):
                tbl[0, col].set_facecolor("#2c3e50")
                tbl[0, col].set_text_props(color="white", fontweight="bold")
            ax.axis("off")
            ax.set_title(f"Summary metrics (N={N})")

    plt.tight_layout()
    _save_figure(fig, "figure_1_algorithm_comparison.pdf", save)
    plt.show()


# ============================================================================
# Section 2 — QSVT circuit complexity analysis
# ============================================================================

def run_section_2() -> dict:
    """
    Compute QSVT circuit complexity metrics analytically for N in
    {4, 8, 16, 32} and compare against HHL circuit depth estimates.

    No quantum circuit simulation is performed in this section — all
    quantities are derived from the condition number scaling and the
    polynomial degree formula d = O(kappa * log(kappa/epsilon)).

    Returns
    -------
    dict
        Complexity metrics for all N values.
    """
    _section_header("QSVT Circuit Complexity Analysis", 2)

    epsilon = 0.01
    results = {}

    print(
        f"\n  {'N':>4}  {'kappa':>8}  {'alpha':>6}  "
        f"{'kappa_eff':>10}  {'deg_QSVT':>10}  "
        f"{'depth_QSVT':>12}  {'depth_HHL_est':>14}  "
        f"{'qubits_QSVT':>12}"
    )
    print(f"  {'─'*82}")

    for N in (4, 8, 16, 32):
        # Condition number: kappa ~ (4/pi^2)(N+1)^2.
        kappa = (4.0 / np.pi**2) * (N + 1)**2

        # Block encoding subnormalisation: alpha = spectral norm of A.
        alpha = subnormalisation_factor(-2.0, 1.0, N)

        # Effective condition number after subnormalisation.
        # kappa_eff = alpha * kappa / ||A||_2
        # For TST with a=-2, b=1: ||A||_2 ~ 4 (max eigenvalue magnitude).
        A_norm_2  = 4.0 * np.cos(np.pi / (N + 1))**0  # approx 4 for large N
        kappa_eff = alpha * kappa / A_norm_2

        # QSVT polynomial degree.
        degree = polynomial_degree_estimate(kappa_eff, epsilon)

        # QSVT circuit depth estimate: degree * (block encoding depth).
        # Block encoding depth for TST: O(n) = O(log N).
        n              = int(np.log2(N))
        be_depth_est   = 10 * n   # empirical estimate from N=4 measurement
        depth_qsvt     = degree * (be_depth_est + 1)

        # HHL circuit depth estimate from Ghafourpour & Laizet (2025) Table V.
        # Hamiltonian simulation: O(sqrt(2^{n_l}) / epsilon * n_b^2)
        # where n_l = ceil(log2(kappa+1)) + 1.
        n_l        = int(np.ceil(np.log2(kappa + 1))) + 1
        depth_hhl  = int(np.sqrt(2**n_l) / epsilon * n**2)

        # QSVT qubit count: n (data) + 2 (block enc. ancilla) + 1 (signal).
        qubits_qsvt = n + 2 + 1

        results[N] = {
            "kappa":       kappa,
            "alpha":       alpha,
            "kappa_eff":   kappa_eff,
            "degree":      degree,
            "depth_qsvt":  depth_qsvt,
            "depth_hhl":   depth_hhl,
            "qubits_qsvt": qubits_qsvt,
            "n":           n,
        }

        print(
            f"  {N:>4}  {kappa:>8.1f}  {alpha:>6.1f}  "
            f"{kappa_eff:>10.1f}  {degree:>10d}  "
            f"{depth_qsvt:>12d}  {depth_hhl:>14d}  "
            f"{qubits_qsvt:>12d}"
        )

    return results


def plot_section_2(data: dict, save: bool = True) -> None:
    """
    Generate Figure 2: QSVT circuit complexity scaling plots.

    Layout (1 row x 3 columns):
        Left:   Polynomial degree vs N for QSVT
        Centre: Circuit depth comparison: QSVT vs HHL
        Right:  Qubit count vs N for QSVT

    Parameters
    ----------
    data : dict
        Output of run_section_2().
    save : bool
    """
    N_vals       = sorted(data.keys())
    kappas       = [data[N]["kappa"]      for N in N_vals]
    degrees      = [data[N]["degree"]     for N in N_vals]
    depths_qsvt  = [data[N]["depth_qsvt"] for N in N_vals]
    depths_hhl   = [data[N]["depth_hhl"]  for N in N_vals]
    qubits       = [data[N]["qubits_qsvt"]for N in N_vals]

    fig, axes = plt.subplots(1, 3, figsize=(14, 5))
    fig.suptitle(
        "Section 2 — QSVT Circuit Complexity Scaling Analysis\n"
        r"1-D Poisson Equation, $\varepsilon = 0.01$",
        fontsize=12,
    )

    # -- Panel 1: polynomial degree vs kappa ----------------------------------
    ax = axes[0]
    ax.loglog(kappas, degrees, color=COLOURS["qsvt"],
              marker="D", ms=8, lw=2, label=r"QSVT degree $d$")
    # Reference line: O(kappa).
    kappa_ref = np.array(kappas, dtype=float)
    ax.loglog(kappa_ref, kappa_ref * degrees[0] / kappas[0],
              color="grey", ls="--", lw=1.2, label=r"$\mathcal{O}(\kappa)$")
    ax.set_xlabel(r"Condition number $\kappa(A)$")
    ax.set_ylabel("Polynomial degree $d$")
    ax.set_title("QSVT polynomial degree")
    ax.legend()
    ax.grid(True, alpha=0.3, which="both")
    for N, kap, deg in zip(N_vals, kappas, degrees):
        ax.annotate(f"N={N}", (kap, deg), textcoords="offset points",
                    xytext=(5, 5), fontsize=8)

    # -- Panel 2: circuit depth comparison ------------------------------------
    ax = axes[1]
    x_pos = np.arange(len(N_vals))
    width = 0.35
    bars1 = ax.bar(x_pos - width/2, depths_qsvt, width,
                   color=COLOURS["qsvt"], alpha=0.85,
                   label="QSVT", edgecolor="black", lw=0.8)
    bars2 = ax.bar(x_pos + width/2, depths_hhl, width,
                   color=COLOURS["hhl"], alpha=0.85,
                   label="HHL (estimate)", edgecolor="black", lw=0.8)
    ax.set_xticks(x_pos)
    ax.set_xticklabels([f"N={N}" for N in N_vals])
    ax.set_ylabel("Circuit depth (gate count)")
    ax.set_title("Circuit depth: QSVT vs HHL")
    ax.set_yscale("log")
    ax.legend()
    ax.grid(True, alpha=0.3, axis="y")

    # -- Panel 3: qubit count -------------------------------------------------
    ax = axes[2]
    ax.plot(N_vals, qubits, color=COLOURS["qsvt"],
            marker="D", ms=8, lw=2, label="QSVT total qubits")
    # HHL qubit count: n_b + n_l + n_MCMT + 1.
    hhl_qubits = []
    for N in N_vals:
        n   = int(np.log2(N))
        kap = data[N]["kappa"]
        n_l = int(np.ceil(np.log2(kap + 1))) + 2
        hhl_qubits.append(n + n_l + (n - 1) + 1)
    ax.plot(N_vals, hhl_qubits, color=COLOURS["hhl"],
            marker="s", ms=8, lw=2, ls="--", label="HHL total qubits")
    ax.set_xlabel("System size N")
    ax.set_ylabel("Total qubit count")
    ax.set_title("Qubit requirements")
    ax.legend()
    ax.grid(True, alpha=0.3)
    for N, q in zip(N_vals, qubits):
        ax.annotate(f"{q}", (N, q), textcoords="offset points",
                    xytext=(4, 4), fontsize=9, color=COLOURS["qsvt"])

    plt.tight_layout()
    _save_figure(fig, "figure_2_qsvt_complexity.pdf", save)
    plt.show()


# ============================================================================
# Section 3 — HET plasma application (1-D)
# ============================================================================

def run_section_3() -> dict:
    """
    Apply HHL and VQLS to the 1-D HET plasma Poisson equation with
    physical parameters from Boeuf & Garrigues (1998).

    Two sub-cases:
        3a: Linear charge density, homogeneous BCs — analytical solution
            available for exact error quantification.
        3b: Gaussian charge density, physical BCs (V_d = 300 V) —
            Thomas serves as the classical reference.

    Returns
    -------
    dict
        Solution vectors, electric field profiles, and error metrics.
    """
    _section_header("HET Plasma Application — 1-D Axial Poisson Equation", 3)

    vqls_cfg = VQLSConfig1D(
        n_layers=6, max_iter=300, tol=1e-6, random_seed=42, verbose=False
    )

    results = {}

    # -- Sub-case 3a: linear profile, homogeneous BCs ------------------------
    print("\n  Sub-case 3a: linear profile, homogeneous BCs (N=8)")
    print(f"  {'Solver':<10} {'MaxRelErr':>10}  {'MaxAbsErr':>12}  "
          f"{'Residual':>12}  {'Time':>8}")
    print(f"  {'─'*60}")

    cfg_a  = HETConfig(N=8, epsilon=0.01, rho_profile="linear",
                       V_discharge=0.0)
    prob_a = HETPoissonProblem1D(cfg_a)
    u_exact_a = HET_EXACT_SOLUTIONS["linear"](
        prob_a.x, cfg_a.rho_0, cfg_a.alpha
    )

    t0 = time.perf_counter()
    u_thomas_a = thomas_solve_system(prob_a.A, prob_a.b)
    t_thomas_a = time.perf_counter() - t0

    t0 = time.perf_counter()
    u_hhl_a, _, _ = hhl_solve_system(prob_a.A, prob_a.b, cfg_a.epsilon)
    t_hhl_a = time.perf_counter() - t0

    t0 = time.perf_counter()
    vr_a   = vqls_solve_system(prob_a.A, prob_a.b, vqls_cfg)
    u_vqls_a = vr_a.u
    t_vqls_a = time.perf_counter() - t0

    for label, u_sol, t_sol, extra in [
        ("Thomas", u_thomas_a, t_thomas_a, ""),
        ("HHL",    u_hhl_a,    t_hhl_a,    ""),
        ("VQLS",   u_vqls_a,   t_vqls_a,   f"cost={vr_a.final_cost:.2e}"),
    ]:
        _solver_row(
            label,
            _max_rel_err(u_sol, u_exact_a),
            float(np.max(np.abs(u_sol - u_exact_a))),
            float(np.linalg.norm(prob_a.A @ u_sol - prob_a.b)
                  / np.linalg.norm(prob_a.b)),
            t_sol, extra,
        )

    results["3a"] = {
        "x": prob_a.x, "cfg": cfg_a,
        "u_exact": u_exact_a,
        "thomas": {"u": u_thomas_a, "t": t_thomas_a},
        "hhl":    {"u": u_hhl_a,    "t": t_hhl_a},
        "vqls":   {"u": u_vqls_a,   "t": t_vqls_a,
                   "cost": vr_a.final_cost},
    }

    # -- Sub-case 3b: Gaussian profile, physical BCs -------------------------
    print("\n  Sub-case 3b: Gaussian profile, V_d=300V (N=8)")
    print(f"  {'Solver':<10} {'MaxRelErr':>10}  {'MaxAbsErr':>12}  "
          f"{'Residual':>12}  {'Time':>8}")
    print(f"  {'─'*60}")

    cfg_b  = HETConfig(N=8, epsilon=0.01, rho_profile="gaussian",
                       V_discharge=300.0)
    prob_b = HETPoissonProblem1D(cfg_b)

    t0 = time.perf_counter()
    u_thomas_b = thomas_solve_system(prob_b.A, prob_b.b)
    t_thomas_b = time.perf_counter() - t0

    t0 = time.perf_counter()
    u_hhl_b, _, _ = hhl_solve_system(prob_b.A, prob_b.b, cfg_b.epsilon)
    t_hhl_b = time.perf_counter() - t0

    t0 = time.perf_counter()
    vr_b   = vqls_solve_system(prob_b.A, prob_b.b, vqls_cfg)
    u_vqls_b = vr_b.u
    t_vqls_b = time.perf_counter() - t0

    ref_b = u_thomas_b
    for label, u_sol, t_sol, extra in [
        ("Thomas", u_thomas_b, t_thomas_b, "(reference)"),
        ("HHL",    u_hhl_b,    t_hhl_b,    ""),
        ("VQLS",   u_vqls_b,   t_vqls_b,   f"cost={vr_b.final_cost:.2e}"),
    ]:
        _solver_row(
            label,
            _max_rel_err(u_sol, ref_b) if label != "Thomas" else 0.0,
            float(np.max(np.abs(u_sol - ref_b))),
            float(np.linalg.norm(prob_b.A @ u_sol - prob_b.b)
                  / np.linalg.norm(prob_b.b)),
            t_sol, extra,
        )

    # Electric field recovery.
    x_full_b, E_thomas_b = _electric_field_1d(
        u_thomas_b, cfg_b.alpha_bc, cfg_b.phi_0, cfg_b.L, cfg_b.N
    )
    _, E_hhl_b  = _electric_field_1d(
        u_hhl_b,  cfg_b.alpha_bc, cfg_b.phi_0, cfg_b.L, cfg_b.N
    )
    _, E_vqls_b = _electric_field_1d(
        u_vqls_b, cfg_b.alpha_bc, cfg_b.phi_0, cfg_b.L, cfg_b.N
    )

    print(f"\n  Peak |E| Thomas: {np.max(np.abs(E_thomas_b)):.3e} V/m")
    print(f"  Peak |E| HHL:    {np.max(np.abs(E_hhl_b)):.3e} V/m")
    print(f"  Peak |E| VQLS:   {np.max(np.abs(E_vqls_b)):.3e} V/m")
    print(f"  B&G (1998) Fig.3 reference: ~2x10^4 V/m near x/L ~ 0.8")

    results["3b"] = {
        "x_int": prob_b.x, "x_full": x_full_b, "cfg": cfg_b,
        "thomas": {"u": u_thomas_b, "E": E_thomas_b, "t": t_thomas_b},
        "hhl":    {"u": u_hhl_b,    "E": E_hhl_b,    "t": t_hhl_b},
        "vqls":   {"u": u_vqls_b,   "E": E_vqls_b,   "t": t_vqls_b,
                   "cost": vr_b.final_cost},
    }

    return results


def plot_section_3(data: dict, save: bool = True) -> None:
    """
    Generate Figure 3: HET plasma 1-D results.

    Layout (2 rows x 3 columns):
        Row 1 (sub-case 3a, linear/homogeneous):
            potential | electric field | relative error
        Row 2 (sub-case 3b, Gaussian/physical):
            potential | electric field | relative error vs Thomas

    Parameters
    ----------
    data : dict
        Output of run_section_3().
    save : bool
    """
    fig = plt.figure(figsize=(15, 9))
    fig.suptitle(
        "Section 3 — HET Plasma Application: 1-D Axial Poisson Equation\n"
        "Physical parameters: Boeuf & Garrigues (1998), "
        "J. Appl. Phys. 84, 3541",
        fontsize=12,
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.42, wspace=0.32)

    # -- Row 1: sub-case 3a ---------------------------------------------------
    d3a  = data["3a"]
    cfg  = d3a["cfg"]
    x    = d3a["x"]
    x_full = np.concatenate([[0.0], x, [1.0]])

    # Potential.
    ax = fig.add_subplot(gs[0, 0])
    ax.plot(x_full, np.concatenate([[0.0], d3a["u_exact"],         [0.0]]),
            color=COLOURS["analytical"], lw=2.5, label="Analytical")
    ax.plot(x_full, np.concatenate([[0.0], d3a["thomas"]["u"],     [0.0]]),
            color=COLOURS["thomas"], ls="--", marker="o", ms=5, label="Thomas")
    ax.plot(x_full, np.concatenate([[0.0], d3a["hhl"]["u"],        [0.0]]),
            color=COLOURS["hhl"],   ls="-.", marker="s", ms=5, label="HHL")
    ax.plot(x_full, np.concatenate([[0.0], d3a["vqls"]["u"],       [0.0]]),
            color=COLOURS["vqls"],  ls=":",  marker="^", ms=5, label="VQLS")
    ax.set_xlabel(r"$\tilde{x} = x/L$")
    ax.set_ylabel(r"$\tilde{\phi}$")
    ax.set_title(
        "Potential: linear profile\n"
        r"$V_d=0$, analytical solution available"
    )
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # Electric field (homogeneous BCs: alpha_bc=0).
    ax = fig.add_subplot(gs[0, 1])
    for key, label in [
        ("thomas", "Thomas"), ("hhl", "HHL"), ("vqls", "VQLS")
    ]:
        xE, E = _electric_field_1d(
            d3a[key]["u"], 0.0, cfg.phi_0, cfg.L, cfg.N
        )
        ax.plot(xE, E / 1e3, color=COLOURS[key],
                ls="--" if key == "thomas" else "-",
                marker=MARKERS[key], ms=5, label=label)
    # Analytical electric field: d/dx[-sin(pi*x)/pi^2] = cos(pi*x)/pi.
    xE_an = np.linspace(0, 1, 100)
    E_an  = (cfg.alpha * cfg.rho_0 / 6.0) * (1.0 - 3.0 * xE_an**2)
    E_an_phys = E_an * cfg.phi_0 / cfg.L
    ax.plot(xE_an, E_an_phys / 1e3, color=COLOURS["analytical"],
            lw=2.5, label="Analytical")
    ax.set_xlabel(r"$\tilde{x} = x/L$")
    ax.set_ylabel(r"$E$ [kV/m]")
    ax.set_title("Electric field: linear profile")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # Relative error.
    ax = fig.add_subplot(gs[0, 2])
    for key, label in [("thomas", "Thomas"), ("hhl", "HHL"), ("vqls", "VQLS")]:
        err = _rel_err_pct(d3a[key]["u"], d3a["u_exact"])
        ax.semilogy(x, err, color=COLOURS[key],
                    ls="--" if key == "thomas" else "-",
                    marker=MARKERS[key], ms=5, label=label)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel("Relative error (%)")
    ax.set_title("Error vs analytical")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3, which="both")

    # -- Row 2: sub-case 3b ---------------------------------------------------
    d3b   = data["3b"]
    cfg_b = d3b["cfg"]
    x_int = d3b["x_int"]
    x_full_b = d3b["x_full"]

    # Potential.
    ax = fig.add_subplot(gs[1, 0])
    phi_full = {
        k: np.concatenate([[cfg_b.alpha_bc], d3b[k]["u"], [0.0]])
        for k in ("thomas", "hhl", "vqls")
    }
    for key, label in [
        ("thomas", "Thomas"), ("hhl", "HHL"), ("vqls", "VQLS")
    ]:
        ax.plot(x_full_b, phi_full[key], color=COLOURS[key],
                ls="--" if key == "thomas" else "-",
                marker=MARKERS[key], ms=4, label=label)
    ax.set_xlabel(r"$\tilde{x} = x/L$")
    ax.set_ylabel(r"$\tilde{\phi}$")
    ax.set_title(
        "Potential: Gaussian profile\n"
        r"$V_d=300$ V, Thomas = reference"
    )
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # Electric field.
    ax = fig.add_subplot(gs[1, 1])
    for key, label in [
        ("thomas", "Thomas"), ("hhl", "HHL"), ("vqls", "VQLS")
    ]:
        ax.plot(x_full_b, d3b[key]["E"] / 1e4, color=COLOURS[key],
                ls="--" if key == "thomas" else "-",
                marker=MARKERS[key], ms=4, label=label)
    ax.axhline(2.0, color="grey", ls="-.", lw=1.2, alpha=0.7,
               label=r"B&G (1998): $\sim 2\times10^4$ V/m")
    ax.axvline(0.8, color="grey", ls=":",  lw=1.0, alpha=0.5)
    ax.set_xlabel(r"$\tilde{x} = x/L$")
    ax.set_ylabel(r"$E$ [$\times 10^4$ V/m]")
    ax.set_title(
        "Electric field: Gaussian profile\n"
        "cf. Boeuf & Garrigues (1998), Fig. 3"
    )
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # Relative error vs Thomas.
    ax = fig.add_subplot(gs[1, 2])
    ref_b = d3b["thomas"]["u"]
    for key, label in [("hhl", "HHL"), ("vqls", "VQLS")]:
        err = _rel_err_pct(d3b[key]["u"], ref_b)
        ax.semilogy(x_int, err, color=COLOURS[key],
                    marker=MARKERS[key], ms=5, label=label)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel("Relative error vs Thomas (%)")
    ax.set_title("Error vs Thomas reference")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3, which="both")

    plt.tight_layout()
    _save_figure(fig, "figure_3_het_1d.pdf", save)
    plt.show()


# ============================================================================
# Section 4 — HET plasma application (2-D)
# ============================================================================

def run_section_4() -> dict:
    """
    Apply Thomas-2D and VQLS-2D to the 2-D HET sinusoidal Poisson
    problem with the analytical solution phi = sin(pi*x)*sin(pi*y).

    Uses N=4 for tractability. The analytical solution provides exact
    error quantification in two dimensions.

    Returns
    -------
    dict
        Solution fields, error arrays, and scalar metrics.
    """
    _section_header(
        "HET Plasma Application — 2-D Poisson, Sinusoidal Source", 4
    )

    cfg     = HETConfig2D(N=4, epsilon=0.01, max_iter=300)
    problem = HETSinusoidalProblem2D(cfg)
    u_exact = problem.analytical_solution()
    Ex_exact, Ey_exact = problem.analytical_electric_field()

    print(f"  {problem.summary()}")
    print(f"  max|phi_exact| = {np.max(np.abs(u_exact)):.4f}  (expected 1.0)")
    print(f"\n  {'Solver':<12} {'Iters':>6}  {'Conv':>5}  "
          f"{'MaxRelErr':>10}  {'MaxAbsErr':>12}  {'Time':>8}")
    print(f"  {'─'*62}")

    inner_cfg = VQLSConfig1D(
        n_layers=3, max_iter=100, tol=1e-2, random_seed=0, verbose=False
    )
    vqls_cfg_2d = VQLSConfig2D(
        inner_config=inner_cfg, warm_start=True, verbose=False
    )

    results = {
        "cfg": cfg, "problem": problem,
        "u_exact": u_exact,
        "Ex_exact": Ex_exact, "Ey_exact": Ey_exact,
    }

    for label, solver_fn, kwargs in [
        ("Thomas-2D", thomas_solve_2d,  {}),
        ("VQLS-2D",   vqls_solve_2d,    {"config": vqls_cfg_2d}),
    ]:
        t0 = time.perf_counter()
        r  = solver_fn(problem, **kwargs)
        t  = time.perf_counter() - t0

        rel = _max_rel_err(r.u.ravel(), u_exact.ravel())
        abs_e = float(np.max(np.abs(r.u - u_exact)))
        conv  = "Yes" if r.converged else "No"
        print(
            f"  {label:<12} {r.iterations:>6}  {conv:>5}  "
            f"{rel:>10.3f}%  {abs_e:>12.4e}  {t:>8.2f}s"
        )
        results[label] = {"result": r, "time": t}

    return results


def plot_section_4(data: dict, save: bool = True) -> None:
    """
    Generate Figure 4: 2-D HET sinusoidal Poisson results.

    Layout (2 rows x 3 columns):
        Row 1: analytical solution | Thomas-2D solution | VQLS-2D solution
        Row 2: Thomas error        | VQLS error         | electric field

    Parameters
    ----------
    data : dict
        Output of run_section_4().
    save : bool
    """
    problem  = data["problem"]
    u_exact  = data["u_exact"]
    X, Y     = problem.X, problem.Y
    r_thomas = data["Thomas-2D"]["result"]
    r_vqls   = data["VQLS-2D"]["result"]

    u_all    = np.stack([u_exact, r_thomas.u, r_vqls.u])
    u_min, u_max = u_all.min(), u_all.max()
    levels_u = np.linspace(u_min, u_max, 25)

    err_thomas = np.abs(r_thomas.u - u_exact)
    err_vqls   = np.abs(r_vqls.u   - u_exact)
    err_max    = max(err_thomas.max(), err_vqls.max())
    levels_e   = np.linspace(0.0, err_max if err_max > 0 else 1.0, 20)

    E_mag_exact = np.sqrt(data["Ex_exact"]**2 + data["Ey_exact"]**2)

    fig = plt.figure(figsize=(15, 9))
    fig.suptitle(
        "Section 4 — HET Plasma Application: 2-D Poisson Equation\n"
        r"Sinusoidal source, analytical solution "
        r"$\tilde{\phi} = \sin(\pi\tilde{x})\sin(\pi\tilde{y})$, N=4",
        fontsize=12,
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.38, wspace=0.32)

    # Row 1: solution contours.
    for col, (Z, title) in enumerate([
        (u_exact,    "Analytical solution"),
        (r_thomas.u, f"Thomas-2D ({r_thomas.iterations} iters)"),
        (r_vqls.u,   f"VQLS-2D ({r_vqls.iterations} iters)"),
    ]):
        ax = fig.add_subplot(gs[0, col])
        cf = ax.contourf(X, Y, Z, levels=levels_u, cmap="viridis")
        ax.contour(X, Y, Z, levels=levels_u,
                   colors="white", linewidths=0.3, alpha=0.4)
        fig.colorbar(cf, ax=ax, shrink=0.85)
        ax.set_xlabel(r"$\tilde{x}$")
        ax.set_ylabel(r"$\tilde{y}$")
        ax.set_title(title)
        ax.set_aspect("equal")

    # Row 2, col 0-1: error contours.
    for col, (Z, title) in enumerate([
        (err_thomas, f"Thomas-2D error\n"
                     f"max={err_thomas.max():.3e}"),
        (err_vqls,   f"VQLS-2D error\n"
                     f"max={err_vqls.max():.3e}"),
    ]):
        ax = fig.add_subplot(gs[1, col])
        cf = ax.contourf(X, Y, Z, levels=levels_e, cmap="hot_r")
        fig.colorbar(cf, ax=ax, shrink=0.85,
                     label=r"$|\tilde{\phi}_{solver} - \tilde{\phi}_{exact}|$")
        ax.set_xlabel(r"$\tilde{x}$")
        ax.set_ylabel(r"$\tilde{y}$")
        ax.set_title(title)
        ax.set_aspect("equal")

    # Row 2, col 2: analytical electric field magnitude.
    ax = fig.add_subplot(gs[1, 2])
    cf = ax.contourf(X, Y, E_mag_exact, levels=20, cmap="plasma")
    ax.quiver(
        X, Y,
        data["Ex_exact"] / E_mag_exact,
        data["Ey_exact"] / E_mag_exact,
        alpha=0.6, scale=20, color="white", width=0.005,
    )
    fig.colorbar(cf, ax=ax, label=r"$|\mathbf{E}|$ [V/m]")
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$\tilde{y}$")
    ax.set_title("Analytical electric field\n(magnitude and direction)")
    ax.set_aspect("equal")

    plt.tight_layout()
    _save_figure(fig, "figure_4_het_2d.pdf", save)
    plt.show()


# ============================================================================
# Section 5 — 2-D QSVT
# ============================================================================

def run_section_5() -> dict:
    """
    Apply QSVT-2D to the 2-D HET sinusoidal Poisson problem.

    The 2-D QSVT solver uses the same line-Jacobi decomposition as
    hhl_2d.py and vqls_2d.py, calling qsvt_solve_system for each row
    sub-problem. The row matrix has a=-4, b=1, kappa_row -> 3, which
    gives a polynomial degree d = O(3 * log(1/epsilon)) ~ constant in N.
    This near-constant degree is the key advantage of QSVT in 2-D.

    Returns
    -------
    dict
        Solution field, error metrics, and circuit diagnostics.
    """
    _section_header(
        "2-D QSVT — HET Sinusoidal Poisson (EXPERIMENTAL)", 5
    )

    from solvers.quantum.qsvt_2d import QSVTConfig2D, qsvt_solve_2d

    cfg     = HETConfig2D(N=4, epsilon=0.01, max_iter=300)
    problem = HETSinusoidalProblem2D(cfg)
    u_exact = problem.analytical_solution()

    # QSVT-2D config: epsilon=0.1 for tractable circuit depth.
    # The row matrix has kappa_row ~ 2.36 (N=4), giving degree ~ 33 via pyqsp.
    # This is still substantially smaller than the 1-D case (degree ~ 45 at N=4).
    qsvt_cfg_2d = QSVTConfig2D(
        epsilon      = 0.1,
        angle_method = "auto",
        max_degree   = 50,
        verbose      = True,
    )

    t0     = time.perf_counter()
    r_qsvt = qsvt_solve_2d(problem, config=qsvt_cfg_2d)
    t_qsvt = time.perf_counter() - t0

    rel = _max_rel_err(r_qsvt.u.ravel(), u_exact.ravel())
    print(
        f"  QSVT-2D: iters={r_qsvt.iterations}, "
        f"converged={r_qsvt.converged}, "
        f"MaxRelErr={rel:.3f}%, time={t_qsvt:.1f}s"
    )

    return {
        "cfg": cfg, "problem": problem,
        "u_exact": u_exact,
        "qsvt": {"result": r_qsvt, "time": t_qsvt},
    }

def plot_section_5(data: dict, s4_data: dict, save: bool = True) -> None:
    """
    Generate Figure 5: 2-D QSVT solution contours, error map, and
    algorithm comparison bar chart.

    Layout (1 row x 3 columns):
        Left:   QSVT-2D solution contour vs analytical
        Centre: Absolute error vs analytical solution
        Right:  Max relative error comparison: Thomas vs VQLS vs QSVT

    Parameters
    ----------
    data : dict
        Output of run_section_5().
    s4_data : dict
        Output of run_section_4(), used for the comparison bar chart.
    save : bool
    """
    problem  = data["problem"]
    u_exact  = data["u_exact"]
    X, Y     = problem.X, problem.Y
    r_qsvt   = data["qsvt"]["result"]

    fig, axes = plt.subplots(1, 3, figsize=(15, 5))
    fig.suptitle(
        r"Section 5 — 2-D QSVT: HET Sinusoidal Poisson, "
        r"$\tilde{\phi} = \sin(\pi\tilde{x})\sin(\pi\tilde{y})$, N=4",
        fontsize=12,
    )

    # -- Panel 1: QSVT solution contour vs analytical overlay -----------------
    u_all    = np.stack([u_exact, r_qsvt.u])
    levels_u = np.linspace(u_all.min(), u_all.max(), 20)

    cf = axes[0].contourf(X, Y, r_qsvt.u, levels=levels_u, cmap="viridis")
    # Overlay analytical contour lines for direct visual comparison.
    axes[0].contour(X, Y, u_exact, levels=8,
                    colors=COLOURS["analytical"], linewidths=0.9,
                    linestyles="--", alpha=0.8)
    fig.colorbar(cf, ax=axes[0])
    axes[0].set_xlabel(r"$\tilde{x}$")
    axes[0].set_ylabel(r"$\tilde{y}$")
    axes[0].set_title(
        f"QSVT-2D solution ({r_qsvt.iterations} iters)\n"
        "Dashed: analytical contours"
    )
    axes[0].set_aspect("equal")

    # -- Panel 2: absolute error vs analytical --------------------------------
    err_qsvt = np.abs(r_qsvt.u - u_exact)
    cf = axes[1].contourf(X, Y, err_qsvt, levels=20, cmap="hot_r")
    fig.colorbar(
        cf, ax=axes[1],
        label=r"$|\tilde{\phi}_\mathrm{QSVT} - \tilde{\phi}_\mathrm{exact}|$",
    )
    axes[1].set_xlabel(r"$\tilde{x}$")
    axes[1].set_ylabel(r"$\tilde{y}$")
    axes[1].set_title(
        f"QSVT-2D absolute error\n"
        f"max = {err_qsvt.max():.3e}"
    )
    axes[1].set_aspect("equal")

    # -- Panel 3: algorithm comparison bar chart ------------------------------
    # Collect max relative errors for all three solvers from Sections 4 and 5.
    u_exact_4 = s4_data["u_exact"]

    def _max_rel_2d(u: np.ndarray, ref: np.ndarray) -> float:
        scale = np.max(np.abs(ref))
        mask  = np.abs(ref) > 1e-4 * scale
        if not mask.any():
            return float("nan")
        return float(np.max(
            np.abs((u - ref)[mask]) / np.abs(ref[mask])
        )) * 100.0

    rel_thomas = _max_rel_2d(
        s4_data["Thomas-2D"]["result"].u, u_exact_4
    )
    rel_vqls   = _max_rel_2d(
        s4_data["VQLS-2D"]["result"].u,   u_exact_4
    )
    rel_qsvt   = _max_rel_2d(r_qsvt.u, u_exact)

    solvers  = ["Thomas-2D", "VQLS-2D", "QSVT-2D"]
    rel_errs = [rel_thomas, rel_vqls, rel_qsvt]
    colours  = [COLOURS["thomas"], COLOURS["vqls"], COLOURS["qsvt"]]

    bars = axes[2].bar(
        solvers, rel_errs,
        color=colours, alpha=0.85,
        edgecolor="black", linewidth=0.8,
    )
    for bar, val in zip(bars, rel_errs):
        axes[2].text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() * 1.02,
            f"{val:.2f}%",
            ha="center", va="bottom", fontsize=10,
        )
    axes[2].set_ylabel("Max relative error (%)")
    axes[2].set_title(
        "Algorithm comparison\n"
        r"2-D HET sinusoidal, $\tilde{\phi}_\mathrm{exact}$ reference"
    )
    axes[2].grid(True, alpha=0.3, axis="y")
    axes[2].set_ylim(0, max(rel_errs) * 1.25)

    plt.tight_layout()
    _save_figure(fig, "figure_5_qsvt_2d.pdf", save)
    plt.show()


# ============================================================================
# Excel / CSV export
# ============================================================================

def export_excel(
    s1_data : dict,
    s2_data : dict,
    s3_data : dict,
    s4_data : dict,
) -> None:
    """
    Export all scalar benchmark metrics to a single Excel workbook with
    one sheet per section.

    Requires openpyxl (pip install openpyxl). Falls back to CSV export
    if openpyxl is not available.

    Parameters
    ----------
    s1_data : dict  Output of run_section_1().
    s2_data : dict  Output of run_section_2().
    s3_data : dict  Output of run_section_3().
    s4_data : dict  Output of run_section_4().
    """
    if _EXCEL_AVAILABLE:
        _export_excel_openpyxl(s1_data, s2_data, s3_data, s4_data)
    else:
        _export_csv_fallback(s1_data, s2_data, s3_data, s4_data)


def _export_excel_openpyxl(
    s1_data : dict,
    s2_data : dict,
    s3_data : dict,
    s4_data : dict,
) -> None:
    """Write results to an Excel workbook using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb   = openpyxl.Workbook()
    path = RESULTS_DIR / "meeting_report_metrics.xlsx"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2C3E50")
    center_align = Alignment(horizontal="center")

    def _write_sheet(ws, headers: list, rows: list) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    # -- Sheet 1: Algorithm comparison (Section 1) ----------------------------
    ws1 = wb.active
    ws1.title = "S1 Algorithm Comparison"
    headers1 = [
        "N", "Solver", "Max Rel. Err. (%)", "Euclidean Residual",
        "Wall Time (s)", "VQLS Cost", "QSVT Degree", "QSVT Depth",
    ]
    rows1 = []
    for N in (4, 8):
        d = s1_data[N]
        for key, label in [
            ("thomas", "Thomas"), ("hhl", "HHL"), ("vqls", "VQLS")
        ]:
            rows1.append([
                N, label,
                round(d[key]["rel"], 4),
                f"{d[key]['res']:.4e}",
                round(d[key]["t"], 3),
                round(d["vqls"]["cost"], 6) if key == "vqls" else "",
                "", "",
            ])
        if d["qsvt"] is not None:
            rows1.append([
                N, "QSVT",
                round(d["qsvt"]["rel"], 4),
                f"{d['qsvt']['res']:.4e}",
                round(d["qsvt"]["t"], 3),
                "",
                d["qsvt"]["degree"],
                d["qsvt"]["depth"],
            ])
        else:
            rows1.append([N, "QSVT", "N/A", "N/A", "N/A", "", "", ""])
    _write_sheet(ws1, headers1, rows1)

    # -- Sheet 2: QSVT complexity (Section 2) ---------------------------------
    ws2 = wb.create_sheet("S2 QSVT Complexity")
    headers2 = [
        "N", "kappa(A)", "alpha", "kappa_eff",
        "QSVT Degree", "QSVT Depth (est.)",
        "HHL Depth (est.)", "QSVT Qubits",
    ]
    rows2 = []
    for N in sorted(s2_data.keys()):
        d = s2_data[N]
        rows2.append([
            N, round(d["kappa"], 1), d["alpha"],
            round(d["kappa_eff"], 1), d["degree"],
            d["depth_qsvt"], d["depth_hhl"], d["qubits_qsvt"],
        ])
    _write_sheet(ws2, headers2, rows2)

    # -- Sheet 3: HET 1-D (Section 3) -----------------------------------------
    ws3 = wb.create_sheet("S3 HET 1D")
    headers3 = [
        "Sub-case", "Solver", "Max Rel. Err. (%)", "Euclidean Residual",
        "Wall Time (s)", "VQLS Cost",
    ]
    rows3 = []
    for sub, label_sub in [("3a", "Linear/Homogeneous"),
                            ("3b", "Gaussian/Physical BCs")]:
        d = s3_data[sub]
        ref = d.get("u_exact", d.get("thomas", {}).get("u"))
        for key, label in [("thomas", "Thomas"), ("hhl", "HHL"),
                            ("vqls", "VQLS")]:
            rel = _max_rel_err(d[key]["u"], ref) if key != "thomas" or sub == "3a" else 0.0
            rows3.append([
                label_sub, label,
                round(rel, 4),
                f"{float(np.linalg.norm(d[key]['u'] - ref) / max(np.linalg.norm(ref), 1e-14)):.4e}",
                round(d[key]["t"], 3),
                round(d[key].get("cost", ""), 6) if key == "vqls" else "",
            ])
    _write_sheet(ws3, headers3, rows3)

    # -- Sheet 4: HET 2-D (Section 4) -----------------------------------------
    ws4 = wb.create_sheet("S4 HET 2D")
    headers4 = [
        "Solver", "Iterations", "Converged",
        "Max Rel. Err. (%)", "Max Abs. Err.", "Wall Time (s)",
    ]
    rows4 = []
    u_exact_4 = s4_data["u_exact"]
    for key, label in [("Thomas-2D", "Thomas-2D"), ("VQLS-2D", "VQLS-2D")]:
        r = s4_data[key]["result"]
        rows4.append([
            label, r.iterations, r.converged,
            round(_max_rel_err(r.u.ravel(), u_exact_4.ravel()), 4),
            f"{float(np.max(np.abs(r.u - u_exact_4))):.4e}",
            round(s4_data[key]["time"], 3),
        ])
    _write_sheet(ws4, headers4, rows4)

    wb.save(path)
    print(f"\n  Excel workbook saved to {path}")


def _export_csv_fallback(
    s1_data : dict,
    s2_data : dict,
    s3_data : dict,
    s4_data : dict,
) -> None:
    """Export results to individual CSV files when openpyxl is unavailable."""
    import csv

    def _write_csv(filename: str, headers: list, rows: list) -> None:
        filepath = RESULTS_DIR / filename
        with open(filepath, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        print(f"  CSV saved to {filepath}")

    # Section 1.
    rows1 = []
    for N in (4, 8):
        d = s1_data[N]
        for key, label in [
            ("thomas", "Thomas"), ("hhl", "HHL"), ("vqls", "VQLS")
        ]:
            rows1.append([
                N, label,
                round(d[key]["rel"], 4),
                f"{d[key]['res']:.4e}",
                round(d[key]["t"], 3),
            ])
    _write_csv(
        "s1_algorithm_comparison.csv",
        ["N", "Solver", "MaxRelErr%", "Residual", "Time_s"],
        rows1,
    )

    # Section 2.
    rows2 = [
        [N, round(s2_data[N]["kappa"], 1), s2_data[N]["degree"],
         s2_data[N]["depth_qsvt"], s2_data[N]["depth_hhl"],
         s2_data[N]["qubits_qsvt"]]
        for N in sorted(s2_data.keys())
    ]
    _write_csv(
        "s2_qsvt_complexity.csv",
        ["N", "kappa", "QSVT_degree", "QSVT_depth", "HHL_depth", "QSVT_qubits"],
        rows2,
    )

    print(f"  CSV files saved to {RESULTS_DIR}")


# ============================================================================
# Utility
# ============================================================================

def _save_figure(fig, filename: str, save: bool) -> None:
    """Save figure as both PDF and PNG to RESULTS_DIR."""
    if not save:
        return
    for ext in ("pdf", "png"):
        path = RESULTS_DIR / filename.replace(".pdf", f".{ext}")
        fig.savefig(path, bbox_inches="tight", dpi=150)
    print(f"  Figure saved to {RESULTS_DIR / filename}")


# ============================================================================
# Main entry point
# ============================================================================

def main() -> None:
    """
    Execute the full meeting progress report.

    Sections are run sequentially. Total estimated runtime: 35-55 minutes
    on a standard laptop (Intel Core i7, 16 GB RAM).

    To reduce runtime for a quick preview:
        - Set QSVTConfig(epsilon=0.2) in run_section_1 to reduce degree
        - Set max_iter=100 in HETConfig2D in run_section_4
        - Comment out run_section_3 if HET 1-D results are already known
    """
    t_start = time.perf_counter()

    print("\n" + "═"*68)
    print("  QUANTUM POISSON SOLVER — MEETING PROGRESS REPORT")
    print("  Imperial College London, Department of Aeronautics")
    print("  HHL | VQLS | QSVT — Poisson Equation and HET Plasma Modelling")
    print("═"*68)
    print(f"  Output directory: {RESULTS_DIR.resolve()}")

    s1_data = run_section_1()
    s2_data = run_section_2()
    s3_data = run_section_3()
    s4_data = run_section_4()
    s5_data = run_section_5()

    t_elapsed = time.perf_counter() - t_start
    print(f"\n{'─'*68}")
    print(f"  All sections completed in {t_elapsed:.1f}s. "
          f"Generating figures and exports...")

    plot_section_1(s1_data, save=True)
    plot_section_2(s2_data, save=True)
    plot_section_3(s3_data, save=True)
    plot_section_4(s4_data, save=True)
    plot_section_5(s5_data, s4_data, save=True)

    export_excel(s1_data, s2_data, s3_data, s4_data)

    print(f"\n  Total elapsed time: {time.perf_counter() - t_start:.1f}s")
    print(f"  All outputs saved to: {RESULTS_DIR.resolve()}")
    print("═"*68)


if __name__ == "__main__":
    main()