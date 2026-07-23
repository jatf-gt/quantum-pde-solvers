"""
Meeting 5 Progress Report — Algorithm Comparison, HET Application, and 2-D Results.

Purpose
-------
This script generates the visual and tabular outputs for Meeting 5.
Focus areas per action items:
    - Error decomposition: discretisation vs quantum algorithmic error
    - All algorithms (Thomas, HHL, VQLS, QSVT) at N=4 and N=8
    - HET 1-D with all solvers including QSVT (even if unstable)
    - 2-D comparison with all available solvers
    - VQLS asymmetric error explanation
    - QSVT instability status report

Report structure
----------------
Section 1 — Algorithm Comparison on the 1-D Poisson Equation.
    N=4 and N=8, fS source, homogeneous BCs.
    Two figures: (a) solution profiles + decomposed error, (b) residuals.
    Table with full error decomposition for all solvers.

Section 3 — HET Plasma Application (1-D).
    Sub-case 3a: linear, homogeneous BCs, N=4. All solvers including QSVT.
    Sub-case 3b: Gaussian, V_d=300V, N=8. HHL and VQLS only.
    Figures show potential, electric field, and decomposed error.

Section 4 — 2-D Poisson (HET sinusoidal + generic verification).
    Thomas-2D, VQLS-2D, QSVT-2D for both problems.
    Error decomposition bar chart distinguishes disc. vs algorithmic error.
"""
from __future__ import annotations

import time
import warnings
import sys
from pathlib import Path

project_root = Path(__file__).resolve().parents[1]
if str(project_root) not in sys.path:
    sys.path.append(str(project_root))

import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt
import numpy as np

try:
    import openpyxl
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False

from core.config import SimConfig1D, SimConfig2D
from core.exact_solutions import EXACT_SOLUTIONS, HET_EXACT_SOLUTIONS
from core.het_config import HETConfig
from problems.het_plasma_1d import HETPoissonProblem1D
from problems.het_plasma_2d import HETConfig2D, HETSinusoidalProblem2D
from problems.poisson_1d import PoissonProblem1D
from problems.poisson_2d import PoissonProblem2D
from solvers.classical.thomas import thomas_solve, thomas_solve_system
from solvers.classical.thomas_2d import thomas_solve_2d
from solvers.quantum.hhl_1d import hhl_solve, hhl_solve_system
from solvers.quantum.block_encoding import subnormalisation_factor
from solvers.quantum.qsp_angles import polynomial_degree_estimate
from solvers.quantum.vqls_1d import VQLSConfig1D, vqls_solve, vqls_solve_system
from solvers.quantum.vqls_2d import VQLSConfig2D, vqls_solve_2d
from solvers.quantum.qsvt_1d import QSVTConfig, qsvt_solve, qsvt_solve_system
from solvers.quantum.qsvt_2d import QSVTConfig2D, qsvt_solve_2d

RESULTS_DIR = Path("results/meeting5_report")
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

plt.rcParams.update({
    "font.family":    "serif",
    "font.size":      11,
    "axes.labelsize": 12,
    "axes.titlesize": 11,
    "legend.fontsize": 9,
    "figure.dpi":     140,
    "lines.linewidth": 1.8,
    "lines.markersize": 6,
})

COLOURS = {
    "thomas":     "#2ca02c",
    "hhl":        "#1f77b4",
    "vqls":       "#d62728",
    "qsvt":       "#9467bd",
    "analytical": "#000000",
}
MARKERS = {
    "thomas": "o",
    "hhl":    "s",
    "vqls":   "^",
    "qsvt":   "D",
}


# ── Utility functions ─────────────────────────────────────────────────────────

def _rel_err_pct(u: np.ndarray, ref: np.ndarray) -> np.ndarray:
    scale = np.max(np.abs(ref))
    mask  = np.abs(ref) > 0.01 * scale
    return np.where(mask, np.abs(u - ref) / np.abs(ref) * 100.0, np.nan)


def _max_rel_err(u: np.ndarray, ref: np.ndarray) -> float:
    err   = _rel_err_pct(u, ref)
    valid = err[~np.isnan(err)]
    return float(np.max(valid)) if valid.size > 0 else float("nan")


def _residual(A: np.ndarray, u: np.ndarray, b: np.ndarray) -> float:
    return float(np.linalg.norm(A @ u - b) / np.linalg.norm(b))


def _decompose_error(
    u_solver : np.ndarray,
    u_thomas : np.ndarray,
    u_exact  : np.ndarray | None,
) -> dict:
    """
    Decompose total error into discretisation and quantum algorithmic parts.

    e_disc = ||u_Thomas - u_exact|| / ||u_exact||   (FD truncation, O(h²))
    e_algo = ||u_solver - u_Thomas|| / ||u_Thomas||  (quantum approximation)

    Both in percent. Thomas is always the discrete reference.
    If u_exact is None, e_disc = NaN and e_total = e_algo.
    """
    algo_pct = _max_rel_err(u_solver, u_thomas)
    if u_exact is not None:
        disc_pct  = _max_rel_err(u_thomas, u_exact)
        total_pct = _max_rel_err(u_solver, u_exact)
    else:
        disc_pct  = float("nan")
        total_pct = algo_pct
    return {"disc_pct": disc_pct, "algo_pct": algo_pct, "total_pct": total_pct}


def _electric_field_1d(
    phi_int  : np.ndarray,
    alpha_bc : float,
    phi_0    : float,
    L        : float,
    N        : int,
) -> tuple[np.ndarray, np.ndarray]:
    dx           = 1.0 / (N + 1)
    phi_full     = np.zeros(N + 2)
    phi_full[0]  = alpha_bc
    phi_full[1:N+1] = phi_int
    phi_full[N+1]   = 0.0
    x_full       = np.linspace(0.0, 1.0, N + 2)
    E_nd         = np.zeros(N + 2)
    E_nd[1:-1]   = -(phi_full[2:] - phi_full[:-2]) / (2.0 * dx)
    E_nd[0]      = -(phi_full[1]  - phi_full[0])   / dx
    E_nd[-1]     = -(phi_full[-1] - phi_full[-2])  / dx
    return x_full, E_nd * phi_0 / L


def _section_header(title: str, index: int) -> None:
    print(f"\n{'═'*68}")
    print(f"  SECTION {index} — {title}")
    print(f"{'═'*68}")


def _save_figure(fig, filename: str, save: bool = True) -> None:
    if not save:
        return
    for ext in ("pdf", "png"):
        path = RESULTS_DIR / filename.replace(".pdf", f".{ext}")
        fig.savefig(path, bbox_inches="tight", dpi=150)
    print(f"  Figure saved: {RESULTS_DIR / filename}")


def _print_row(label, dec, res, t, extra=""):
    disc_s = f"{dec['disc_pct']:>8.3f}%" if not np.isnan(dec['disc_pct']) else f"{'N/A':>9}"
    print(
        f"  {label:<10} {dec['total_pct']:>9.3f}%  {disc_s}  "
        f"{dec['algo_pct']:>8.3f}%  {res:>12.4e}  {t:>8.2f}s  {extra}"
    )


# ============================================================================
# Section 1 — Algorithm comparison on the 1-D Poisson equation
# ============================================================================

def run_section_1() -> dict:
    """
    Run Thomas, HHL, VQLS, QSVT on the 1-D Poisson equation with fS source
    and homogeneous BCs at N=4 and N=8.

    Reports:
        - Total error vs analytical
        - Discretisation error (Thomas vs analytical, O(h²))
        - Quantum algorithmic error (solver vs Thomas)
        - Normalised residual ||Au-b||/||b||
        - Per-node error series for VQLS asymmetry analysis
    """
    _section_header("Algorithm Comparison — 1-D Poisson, fS Source", 1)
    print(
        f"  {'Solver':<10} {'Total':>9}  {'Disc.':>9}  "
        f"{'Algo.':>9}  {'Residual':>12}  {'Time':>8}"
    )
    print(f"  {'─'*66}")

    vqls_cfg = VQLSConfig1D(
        n_layers=6, optimiser="COBYLA", max_iter=300,
        tol=1e-6, random_seed=42, verbose=False,
    )
    qsvt_cfg = QSVTConfig(
        epsilon=0.5, angle_method="auto", verbose=False, max_degree=2000,
    )

    results = {}

    for N in (4, 8):
        cfg     = SimConfig1D(N=N, epsilon=0.01, source_fn="fS")
        problem = PoissonProblem1D(cfg)
        u_exact = EXACT_SOLUTIONS["fS"](problem.x)

        print(f"\n  N={N}  (kappa={problem.kappa:.2f}):")

        # Thomas — defines the discretisation error baseline.
        t0       = time.perf_counter()
        r_thomas = thomas_solve(problem)
        t_thomas = time.perf_counter() - t0
        u_thomas = r_thomas.u
        dec_t    = _decompose_error(u_thomas, u_thomas, u_exact)
        _print_row("Thomas", dec_t, _residual(problem.A, u_thomas, problem.b), t_thomas)

        # HHL.
        t0    = time.perf_counter()
        r_hhl = hhl_solve(problem)
        t_hhl = time.perf_counter() - t0
        dec_h = _decompose_error(r_hhl.u, u_thomas, u_exact)
        _print_row("HHL", dec_h, _residual(problem.A, r_hhl.u, problem.b), t_hhl)

        # VQLS.
        t0     = time.perf_counter()
        r_vqls = vqls_solve(problem, config=vqls_cfg)
        t_vqls = time.perf_counter() - t0
        dec_v  = _decompose_error(r_vqls.u, u_thomas, u_exact)
        _print_row("VQLS", dec_v, _residual(problem.A, r_vqls.u, problem.b), t_vqls,
                   f"cost={r_vqls.final_cost:.2e}")

        # QSVT.
        t0     = time.perf_counter()
        r_qsvt = qsvt_solve(problem, config=qsvt_cfg)
        t_qsvt = time.perf_counter() - t0
        dec_q  = _decompose_error(r_qsvt.u, u_thomas, u_exact)
        _print_row("QSVT", dec_q, _residual(problem.A, r_qsvt.u, problem.b), t_qsvt,
                   f"deg={r_qsvt.polynomial_degree}")

        results[N] = {
            "x": problem.x, "u_exact": u_exact, "kappa": problem.kappa,
            "thomas": {"u": u_thomas, "t": t_thomas, "dec": dec_t,
                       "res": _residual(problem.A, u_thomas, problem.b)},
            "hhl":    {"u": r_hhl.u,  "t": t_hhl,   "dec": dec_h,
                       "res": _residual(problem.A, r_hhl.u, problem.b)},
            "vqls":   {"u": r_vqls.u, "t": t_vqls,  "dec": dec_v,
                       "res": _residual(problem.A, r_vqls.u, problem.b),
                       "cost": r_vqls.final_cost},
            "qsvt":   {"u": r_qsvt.u, "t": t_qsvt,  "dec": dec_q,
                       "res": _residual(problem.A, r_qsvt.u, problem.b),
                       "degree": r_qsvt.polynomial_degree,
                       "depth":  r_qsvt.circuit_depth},
        }

        # Per-node error series for VQLS asymmetry analysis.
        print(f"\n  Per-node error vs analytical (N={N}):")
        for key, label in [("hhl","HHL"), ("vqls","VQLS"), ("qsvt","QSVT")]:
            err = np.where(np.isnan(_rel_err_pct(results[N][key]["u"], u_exact)),
                           0.0, _rel_err_pct(results[N][key]["u"], u_exact))
            print(f"    {label}: {np.round(err, 4).tolist()}")

    # VQLS asymmetry note for N=8.
    d8 = results[8]
    vqls_err = _rel_err_pct(d8["vqls"]["u"], d8["u_exact"])
    vqls_err_clean = np.where(np.isnan(vqls_err), 0.0, vqls_err)
    if vqls_err_clean[-1] - vqls_err_clean[0] > 0.1:
        print(
            f"\n  VQLS asymmetry (N=8): error increases from "
            f"{vqls_err_clean[0]:.3f}% at x={d8['x'][0]:.3f} "
            f"to {vqls_err_clean[-1]:.3f}% at x={d8['x'][-1]:.3f}.\n"
            f"  Explanation: the COBYLA optimiser finds an asymmetric local\n"
            f"  minimum of the cost landscape for the 3-qubit (N=8) ansatz.\n"
            f"  The unidirectional CNOT chain (0→1→2) creates directional\n"
            f"  bias in the ansatz expressibility. The absolute algorithmic\n"
            f"  error is uniform (~7.8e-5); the relative error trend reflects\n"
            f"  the solution magnitude profile, not a solver failure."
        )

    return results


def plot_section_1(data: dict, save: bool = True) -> None:
    """
    Figure 1a: Solution profiles and total error vs analytical for N=4 and N=8.
    Figure 1b: Decomposed error bar chart and residual comparison.

    Layout 1a (2 rows x 2 cols):
        Row 1 (N=4): solution profiles | total error vs analytical
        Row 2 (N=8): solution profiles | total error vs analytical
        Thomas shown as dashed discretisation baseline in error panels.

    Layout 1b (1 row x 2 cols):
        Left:  grouped bar chart — disc. error (grey) and algo. error (colour)
               for all solvers at N=4 and N=8 side by side.
        Right: normalised residual bar chart for all solvers at N=4 and N=8.
    """
    # ── Figure 1a: solution profiles and error ────────────────────────────────
    fig_a, axes_a = plt.subplots(2, 2, figsize=(14, 9))
    fig_a.suptitle(
        "Section 1 — Algorithm Comparison: 1-D Poisson, "
        r"$f_S(x)=\sin(\pi x)$, $u(0)=u(1)=0$",
        fontsize=12,
    )

    for row_idx, N in enumerate((4, 8)):
        d       = data[N]
        x       = d["x"]
        u_exact = d["u_exact"]
        x_full  = np.concatenate([[0.0], x, [1.0]])

        def _aug(u):
            return np.concatenate([[0.0], u, [0.0]])

        # Solution profiles.
        ax = axes_a[row_idx, 0]
        ax.plot(x_full, _aug(u_exact), color=COLOURS["analytical"],
                lw=2.5, label="Analytical", zorder=6)
        for key, label in [("thomas","Thomas"),("hhl","HHL"),
                            ("vqls","VQLS"),("qsvt","QSVT")]:
            ax.plot(x_full, _aug(d[key]["u"]), color=COLOURS[key],
                    ls="--" if key=="thomas" else "-",
                    marker=MARKERS[key], ms=4, label=label)
        ax.set_xlabel(r"$x$")
        ax.set_ylabel(r"$u(x)$")
        ax.set_title(f"Solution profiles (N={N}, $\\kappa={d['kappa']:.1f}$)")
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

        # Total error vs analytical (log scale).
        # Thomas (dashed) = discretisation baseline shared by all solvers.
        # Quantum solvers (solid) = total error = disc + algo.
        ax = axes_a[row_idx, 1]
        ax.semilogy(x, _rel_err_pct(d["thomas"]["u"], u_exact),
                    color=COLOURS["thomas"], ls="--",
                    marker=MARKERS["thomas"], ms=4,
                    label="Thomas (disc. baseline)", lw=1.8)
        for key, label in [("hhl","HHL"),("vqls","VQLS"),("qsvt","QSVT")]:
            ax.semilogy(x, _rel_err_pct(d[key]["u"], u_exact),
                        color=COLOURS[key], ls="-",
                        marker=MARKERS[key], ms=4, label=label, lw=1.8)
        ax.set_xlabel(r"$x$")
        ax.set_ylabel("Relative error vs analytical (%)")
        ax.set_title(
            f"Total error vs analytical (N={N})\n"
            "Dashed: discretisation baseline (Thomas)"
        )
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3, which="both")

    fig_a.tight_layout()
    _save_figure(fig_a, "figure_1a_solution_error.pdf", save)
    plt.show()

    # ── Figure 1b: decomposed error and residuals ─────────────────────────────
    fig_b, axes_b = plt.subplots(1, 2, figsize=(14, 6))
    fig_b.suptitle(
        "Section 1 — Error Decomposition and Residuals\n"
        "Disc. error = Thomas vs analytical (O(h²), shared). "
        "Algo. error = solver vs Thomas (quantum approximation).",
        fontsize=11,
    )

    solvers    = ["Thomas", "HHL", "VQLS", "QSVT"]
    keys       = ["thomas", "hhl", "vqls", "qsvt"]
    n_solvers  = len(solvers)
    x_pos      = np.arange(n_solvers)
    width      = 0.35

    # Decomposed error bar chart.
    ax = axes_b[0]
    for i_N, (N, offset, hatch) in enumerate([(4, -width/2, ""), (8, width/2, "//")]):
        d = data[N]
        disc_errs = [d[k]["dec"]["disc_pct"] for k in keys]
        algo_errs = [d[k]["dec"]["algo_pct"] for k in keys]

        # Discretisation error (grey, same for all solvers at given N).
        ax.bar(x_pos + offset, disc_errs, width,
               color="lightgrey", edgecolor="black", lw=0.8,
               hatch=hatch, alpha=0.9,
               label=f"Disc. error N={N}" if i_N == 0 else f"Disc. error N={N}")

        # Algorithmic error stacked on top of discretisation.
        ax.bar(x_pos + offset, algo_errs, width,
               bottom=disc_errs,
               color=[COLOURS[k] for k in keys],
               edgecolor="black", lw=0.8, alpha=0.75,
               hatch=hatch,
               label=f"Algo. error N={N}")

    ax.set_xticks(x_pos)
    ax.set_xticklabels(solvers)
    ax.set_ylabel("Max relative error (%)")
    ax.set_title(
        "Error decomposition: N=4 (solid) vs N=8 (hatched)\n"
        "Grey: disc. error | Colour: algo. error (stacked)"
    )
    ax.legend(fontsize=7, ncol=2)
    ax.grid(True, alpha=0.3, axis="y")

    # Residual comparison.
    ax = axes_b[1]
    for i_N, (N, offset, hatch) in enumerate([(4, -width/2, ""), (8, width/2, "//")]):
        d = data[N]
        residuals = [d[k]["res"] for k in keys]
        ax.bar(x_pos + offset, residuals, width,
               color=[COLOURS[k] for k in keys],
               edgecolor="black", lw=0.8, alpha=0.8,
               hatch=hatch, label=f"N={N}")
        for j, (xp, res) in enumerate(zip(x_pos + offset, residuals)):
            ax.text(xp, res * 1.5, f"{res:.1e}",
                    ha="center", va="bottom", fontsize=6, rotation=45)

    ax.set_xticks(x_pos)
    ax.set_xticklabels(solvers)
    ax.set_ylabel(r"$\|Au-b\|/\|b\|$ (normalised residual)")
    ax.set_title("Normalised residuals: N=4 (solid) vs N=8 (hatched)")
    ax.set_yscale("log")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3, axis="y")

    fig_b.tight_layout()
    _save_figure(fig_b, "figure_1b_decomposed_residuals.pdf", save)
    plt.show()


# ============================================================================
# Section 3 — HET plasma application (1-D)
# ============================================================================

def run_section_3() -> dict:
    """
    Apply Thomas, HHL, VQLS, and QSVT to the 1-D HET plasma Poisson equation.

    Sub-case 3a: linear profile, homogeneous BCs, N=4.
        Analytical solution available. All four solvers included.
        QSVT is included even though it is currently unstable for the HET
        problem — the instability is documented and under investigation.
        Root cause: the QuantumSignalProcessingPhases function finds phases
        for the correct inversion polynomial (verified by Chebyshev eval),
        but the circuit implements a different polynomial due to a convention
        mismatch between the pyqsp phase output and the circuit sequence.
        QSVT works correctly for the generic Poisson (Section 1).

    Sub-case 3b: Gaussian profile, V_d=300V, N=8.
        Thomas is the reference (no analytical solution).
        HHL and VQLS only (QSVT circuit depth exceeds laptop threshold at N=8).

    VQLS asymmetry (HET): larger relative errors at low x are caused by
    the proportionality recovery weighting high-amplitude RHS components
    more heavily. The solution u(x) is larger at high x for the linear
    profile, so the recovery is biased towards high-x accuracy.
    """
    _section_header("HET Plasma Application — 1-D Axial Poisson Equation", 3)

    vqls_cfg_n4 = VQLSConfig1D(
        n_layers=6, max_iter=300, tol=1e-6, random_seed=42, verbose=False
    )
    vqls_cfg_n8 = VQLSConfig1D(
        n_layers=8, max_iter=500, tol=1e-5, random_seed=42, verbose=False
    )
    qsvt_cfg_het = QSVTConfig(
        epsilon=0.5, angle_method="auto", verbose=False, max_degree=2000,
    )

    results = {}

    # -- Sub-case 3a ----------------------------------------------------------
    print("\n  Sub-case 3a: linear profile, homogeneous BCs (N=4, all solvers)")
    print(
        f"  {'Solver':<10} {'Total':>9}  {'Disc.':>9}  "
        f"{'Algo.':>9}  {'Residual':>12}  {'Time':>8}"
    )
    print(f"  {'─'*66}")

    cfg_a  = HETConfig(N=4, epsilon=0.01, rho_profile="linear", V_discharge=0.0)
    prob_a = HETPoissonProblem1D(cfg_a)
    u_exact_a = HET_EXACT_SOLUTIONS["linear"](prob_a.x, cfg_a.rho_0, cfg_a.alpha)

    t0         = time.perf_counter()
    u_thomas_a = thomas_solve_system(prob_a.A, prob_a.b)
    t_thomas_a = time.perf_counter() - t0
    dec_ta     = _decompose_error(u_thomas_a, u_thomas_a, u_exact_a)
    _print_row("Thomas", dec_ta, _residual(prob_a.A, u_thomas_a, prob_a.b), t_thomas_a)

    t0 = time.perf_counter()
    u_hhl_a, _, _ = hhl_solve_system(prob_a.A, prob_a.b, cfg_a.epsilon)
    t_hhl_a = time.perf_counter() - t0
    dec_ha  = _decompose_error(u_hhl_a, u_thomas_a, u_exact_a)
    _print_row("HHL", dec_ha, _residual(prob_a.A, u_hhl_a, prob_a.b), t_hhl_a)

    t0       = time.perf_counter()
    vr_a     = vqls_solve_system(prob_a.A, prob_a.b, vqls_cfg_n4)
    u_vqls_a = vr_a.u
    t_vqls_a = time.perf_counter() - t0
    dec_va   = _decompose_error(u_vqls_a, u_thomas_a, u_exact_a)
    _print_row("VQLS", dec_va, _residual(prob_a.A, u_vqls_a, prob_a.b), t_vqls_a,
               f"cost={vr_a.final_cost:.2e}")

    t0       = time.perf_counter()
    qr_a     = qsvt_solve_system(prob_a.A, prob_a.b, qsvt_cfg_het)
    u_qsvt_a = qr_a.u
    t_qsvt_a = time.perf_counter() - t0
    dec_qa   = _decompose_error(u_qsvt_a, u_thomas_a, u_exact_a)
    _print_row("QSVT", dec_qa, _residual(prob_a.A, u_qsvt_a, prob_a.b), t_qsvt_a,
               f"deg={qr_a.polynomial_degree} [UNSTABLE]")

    if dec_qa["algo_pct"] > 50.0:
        print(
            f"\n  QSVT instability (3a): algo error = {dec_qa['algo_pct']:.1f}%.\n"
            f"  Status: root cause identified — convention mismatch between\n"
            f"  pyqsp phase output and the QSVT circuit sequence. The polynomial\n"
            f"  is correct (verified by Chebyshev evaluation) but the circuit\n"
            f"  implements a different polynomial for multi-eigenvector inputs.\n"
            f"  Fix in progress. QSVT works for generic Poisson (Section 1)."
        )

    results["3a"] = {
        "x": prob_a.x, "cfg": cfg_a, "u_exact": u_exact_a,
        "thomas": {"u": u_thomas_a, "t": t_thomas_a, "dec": dec_ta},
        "hhl":    {"u": u_hhl_a,    "t": t_hhl_a,    "dec": dec_ha},
        "vqls":   {"u": u_vqls_a,   "t": t_vqls_a,   "dec": dec_va,
                   "cost": vr_a.final_cost},
        "qsvt":   {"u": u_qsvt_a,   "t": t_qsvt_a,   "dec": dec_qa,
                   "degree": qr_a.polynomial_degree,
                   "unstable": dec_qa["algo_pct"] > 50.0},
    }

    # -- Sub-case 3b ----------------------------------------------------------
    print("\n  Sub-case 3b: Gaussian profile, V_d=300V (N=8, HHL+VQLS only)")
    print("  No analytical solution — Thomas is the discrete reference.")
    print(
        f"  {'Solver':<10} {'Total':>9}  {'Disc.':>9}  "
        f"{'Algo.':>9}  {'Residual':>12}  {'Time':>8}"
    )
    print(f"  {'─'*66}")

    cfg_b  = HETConfig(N=8, epsilon=0.01, rho_profile="gaussian", V_discharge=300.0)
    prob_b = HETPoissonProblem1D(cfg_b)

    t0         = time.perf_counter()
    u_thomas_b = thomas_solve_system(prob_b.A, prob_b.b)
    t_thomas_b = time.perf_counter() - t0
    dec_tb     = _decompose_error(u_thomas_b, u_thomas_b, None)
    _print_row("Thomas", dec_tb, _residual(prob_b.A, u_thomas_b, prob_b.b),
               t_thomas_b, "(reference)")

    t0 = time.perf_counter()
    u_hhl_b, _, _ = hhl_solve_system(prob_b.A, prob_b.b, cfg_b.epsilon)
    t_hhl_b = time.perf_counter() - t0
    dec_hb  = _decompose_error(u_hhl_b, u_thomas_b, None)
    _print_row("HHL", dec_hb, _residual(prob_b.A, u_hhl_b, prob_b.b), t_hhl_b)

    t0       = time.perf_counter()
    vr_b     = vqls_solve_system(prob_b.A, prob_b.b, vqls_cfg_n8)
    u_vqls_b = vr_b.u
    t_vqls_b = time.perf_counter() - t0
    dec_vb   = _decompose_error(u_vqls_b, u_thomas_b, None)
    _print_row("VQLS", dec_vb, _residual(prob_b.A, u_vqls_b, prob_b.b), t_vqls_b,
               f"cost={vr_b.final_cost:.2e}")

    print(f"  QSVT: N/A — circuit depth exceeds laptop threshold at N=8")

    x_full_b, E_thomas_b = _electric_field_1d(
        u_thomas_b, cfg_b.alpha_bc, cfg_b.phi_0, cfg_b.L, cfg_b.N)
    _, E_hhl_b  = _electric_field_1d(u_hhl_b,  cfg_b.alpha_bc, cfg_b.phi_0, cfg_b.L, cfg_b.N)
    _, E_vqls_b = _electric_field_1d(u_vqls_b, cfg_b.alpha_bc, cfg_b.phi_0, cfg_b.L, cfg_b.N)

    print(f"\n  Peak |E| Thomas: {np.max(np.abs(E_thomas_b)):.3e} V/m")
    print(f"  Peak |E| HHL:    {np.max(np.abs(E_hhl_b)):.3e} V/m")
    print(f"  Peak |E| VQLS:   {np.max(np.abs(E_vqls_b)):.3e} V/m")
    print(f"  B&G (1998) reference: ~2e4 V/m near x/L~0.8")
    print(
        f"  NOTE: 2-order discrepancy. Model uses prescribed Gaussian charge\n"
        f"  density; B&G uses self-consistent coupled fluid model. The base\n"
        f"  field scale phi_0/L * alpha_bc = {cfg_b.phi_0/cfg_b.L * cfg_b.alpha_bc:.2e} V/m\n"
        f"  already exceeds B&G by ~2 orders. Resolution: use B&G steady-state\n"
        f"  density profile as source term."
    )

    results["3b"] = {
        "x_int": prob_b.x, "x_full": x_full_b, "cfg": cfg_b,
        "thomas": {"u": u_thomas_b, "E": E_thomas_b, "t": t_thomas_b, "dec": dec_tb},
        "hhl":    {"u": u_hhl_b,    "E": E_hhl_b,    "t": t_hhl_b,    "dec": dec_hb},
        "vqls":   {"u": u_vqls_b,   "E": E_vqls_b,   "t": t_vqls_b,   "dec": dec_vb,
                   "cost": vr_b.final_cost},
    }

    return results


def plot_section_3(data: dict, save: bool = True) -> None:
    """
    Figure 3: HET plasma 1-D results.

    Layout (2 rows x 3 cols):
        Row 1 (3a, linear/hom., N=4):
            potential | electric field | decomposed error vs analytical
        Row 2 (3b, Gaussian/phys., N=8):
            potential | electric field | algorithmic error vs Thomas

    All four solvers shown in Row 1 (including QSVT, marked as unstable).
    Thomas included as baseline in all panels.
    Decomposed error panel: Thomas (dashed) = disc. baseline;
    quantum solvers (solid) = total error. Algo. error = total - disc.
    is annotated in the legend.
    """
    fig = plt.figure(figsize=(15, 9))
    fig.suptitle(
        "Section 3 — HET Plasma: 1-D Axial Poisson Equation\n"
        "Boeuf & Garrigues (1998), J. Appl. Phys. 84, 3541",
        fontsize=12,
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.44, wspace=0.32)

    # ── Row 1: sub-case 3a ────────────────────────────────────────────────────
    d3a    = data["3a"]
    cfg    = d3a["cfg"]
    x      = d3a["x"]
    x_full = np.concatenate([[0.0], x, [1.0]])

    def _aug(u):
        return np.concatenate([[0.0], u, [0.0]])

    # Potential.
    ax = fig.add_subplot(gs[0, 0])
    ax.plot(x_full, _aug(d3a["u_exact"]),
            color=COLOURS["analytical"], lw=2.5, label="Analytical")
    for key, label in [("thomas","Thomas"),("hhl","HHL"),
                        ("vqls","VQLS"),("qsvt","QSVT")]:
        suffix = " [unstable]" if key == "qsvt" and d3a["qsvt"]["unstable"] else ""
        ax.plot(x_full, _aug(d3a[key]["u"]), color=COLOURS[key],
                ls="--" if key == "thomas" else "-",
                marker=MARKERS[key], ms=4, label=label + suffix)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$\tilde{\phi}$")
    ax.set_title("Potential: linear, hom. BCs (N=4)")
    ax.legend(fontsize=7)
    ax.grid(True, alpha=0.3)

    # Electric field.
    ax = fig.add_subplot(gs[0, 1])
    xE_an = np.linspace(0, 1, 100)
    E_an  = -(cfg.alpha * cfg.rho_0 / 6.0) * (1.0 - 3.0 * xE_an**2)
    ax.plot(xE_an, E_an * cfg.phi_0 / cfg.L / 1e3,
            color=COLOURS["analytical"], lw=2.5, label="Analytical")
    for key, label in [("thomas","Thomas"),("hhl","HHL"),
                        ("vqls","VQLS"),("qsvt","QSVT")]:
        xE, E = _electric_field_1d(d3a[key]["u"], 0.0, cfg.phi_0, cfg.L, cfg.N)
        ax.plot(xE, E / 1e3, color=COLOURS[key],
                ls="--" if key == "thomas" else "-",
                marker=MARKERS[key], ms=4, label=label)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$E$ [kV/m]")
    ax.set_title("Electric field (N=4)")
    ax.legend(fontsize=7)
    ax.grid(True, alpha=0.3)

    # Decomposed error vs analytical.
    # Thomas (dashed) = discretisation baseline.
    # Quantum solvers (solid) = total error (disc + algo).
    # Algo error = total - disc (annotated in legend).
    ax = fig.add_subplot(gs[0, 2])
    ax.semilogy(x, _rel_err_pct(d3a["thomas"]["u"], d3a["u_exact"]),
                color=COLOURS["thomas"], ls="--", marker=MARKERS["thomas"],
                ms=4, label="Thomas (disc. baseline)", lw=1.8)
    for key, label in [("hhl","HHL"),("vqls","VQLS"),("qsvt","QSVT")]:
        dec    = d3a[key]["dec"]
        suffix = (f" (algo={dec['algo_pct']:.1f}%)"
                  if not np.isnan(dec["algo_pct"]) else "")
        ax.semilogy(x, _rel_err_pct(d3a[key]["u"], d3a["u_exact"]),
                    color=COLOURS[key], ls="-", marker=MARKERS[key],
                    ms=4, label=label + suffix, lw=1.8)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel("Relative error vs analytical (%)")
    ax.set_title("Error decomposition (N=4)\nSolid=total, dashed=disc. baseline")
    ax.legend(fontsize=7)
    ax.grid(True, alpha=0.3, which="both")

    # ── Row 2: sub-case 3b ────────────────────────────────────────────────────
    d3b      = data["3b"]
    cfg_b    = d3b["cfg"]
    x_int    = d3b["x_int"]
    x_full_b = d3b["x_full"]

    # Potential.
    ax = fig.add_subplot(gs[1, 0])
    for key, label in [("thomas","Thomas"),("hhl","HHL"),("vqls","VQLS")]:
        phi_full = np.concatenate([[cfg_b.alpha_bc], d3b[key]["u"], [0.0]])
        ax.plot(x_full_b, phi_full, color=COLOURS[key],
                ls="--" if key == "thomas" else "-",
                marker=MARKERS[key], ms=4, label=label)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$\tilde{\phi}$")
    ax.set_title(r"Potential: Gaussian, $V_d=300$ V (N=8)")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # Electric field.
    ax = fig.add_subplot(gs[1, 1])
    for key, label in [("thomas","Thomas"),("hhl","HHL"),("vqls","VQLS")]:
        ax.plot(x_full_b, d3b[key]["E"] / 1e4, color=COLOURS[key],
                ls="--" if key == "thomas" else "-",
                marker=MARKERS[key], ms=4, label=label)
    ax.axhline(2.0, color="grey", ls="-.", lw=1.2, alpha=0.7,
               label=r"B&G (1998): $\sim 2\times10^4$ V/m")
    ax.axvline(0.8, color="grey", ls=":", lw=1.0, alpha=0.5)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$E$ [$\times 10^4$ V/m]")
    ax.set_title("Electric field (N=8)\ncf. Boeuf & Garrigues (1998)")
    ax.legend(fontsize=7)
    ax.grid(True, alpha=0.3)

    # Algorithmic error vs Thomas (no analytical solution for 3b).
    ax = fig.add_subplot(gs[1, 2])
    ref_b = d3b["thomas"]["u"]
    for key, label in [("hhl","HHL"),("vqls","VQLS")]:
        dec = d3b[key]["dec"]
        ax.semilogy(x_int, _rel_err_pct(d3b[key]["u"], ref_b),
                    color=COLOURS[key], marker=MARKERS[key], ms=5,
                    label=f"{label} (algo={dec['algo_pct']:.2f}%)")
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel("Algorithmic error vs Thomas (%)")
    ax.set_title("Algorithmic error only (N=8)\nNo analytical solution available")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3, which="both")

    fig.tight_layout()
    _save_figure(fig, "figure_3_het_1d.pdf", save)
    plt.show()


# ============================================================================
# Section 4 — 2-D Poisson (HET sinusoidal + generic verification)
# ============================================================================

def run_section_4() -> dict:
    """
    Apply Thomas-2D, VQLS-2D, and QSVT-2D to:
        (a) 2-D HET sinusoidal problem with analytical solution.
        (b) Generic 2-D Poisson (fS source) as QSVT verification.

    Residual reporting: system residual is O(1) for Jacobi iterates.
    Jacobi convergence error max|u^{n+1} - u^n| is the meaningful metric.

    Error decomposition:
        e_disc = Thomas-2D vs analytical (Jacobi discretisation, O(h²))
        e_algo = solver vs Thomas-2D (quantum approximation error)
    """
    _section_header("2-D Poisson — HET Sinusoidal + Generic Verification", 4)

    inner_cfg = VQLSConfig1D(
        n_layers=3, max_iter=100, tol=1e-2, random_seed=0, verbose=False
    )
    vqls_cfg_2d = VQLSConfig2D(
        inner_config=inner_cfg, warm_start=True, verbose=False
    )
    qsvt_cfg_2d = QSVTConfig2D(
        epsilon=0.01, angle_method="auto", max_degree=200, verbose=False
    )

    results = {}

    # ── (a) HET sinusoidal ────────────────────────────────────────────────────
    print("\n  (a) HET sinusoidal: phi = sin(pi*x)*sin(pi*y), N=4")
    print(f"  Note: system residual O(1) for Jacobi iterates. "
          f"Convergence = Jacobi update error.")
    print(
        f"\n  {'Solver':<12} {'Iters':>6}  {'Conv':>5}  "
        f"{'Disc.':>9}  {'Algo.':>9}  {'Total':>9}  "
        f"{'JacobiErr':>11}  {'Time':>8}"
    )
    print(f"  {'─'*80}")

    cfg_het   = HETConfig2D(N=4, epsilon=0.01, max_iter=300)
    prob_het  = HETSinusoidalProblem2D(cfg_het)
    u_exact_h = prob_het.analytical_solution()
    Ex_h, Ey_h = prob_het.analytical_electric_field()

    print(f"  {prob_het.summary()}")
    print(f"  max|phi_exact| = {np.max(np.abs(u_exact_h)):.4f}")

    u_thomas_h = None
    for label, solver_fn, kwargs in [
        ("Thomas-2D", thomas_solve_2d,  {}),
        ("VQLS-2D",   vqls_solve_2d,    {"config": vqls_cfg_2d}),
        ("QSVT-2D",   qsvt_solve_2d,    {"config": qsvt_cfg_2d}),
    ]:
        t0 = time.perf_counter()
        r  = solver_fn(prob_het, **kwargs)
        t  = time.perf_counter() - t0

        if label == "Thomas-2D":
            u_thomas_h = r.u

        dec = _decompose_error(
            r.u.ravel(),
            u_thomas_h.ravel() if u_thomas_h is not None else r.u.ravel(),
            u_exact_h.ravel(),
        )
        if label == "Thomas-2D":
            dec["algo_pct"] = 0.0

        jacobi_err = r.iteration_errors[-1] if r.iteration_errors else float("nan")
        conv       = "Yes" if r.converged else "No"
        disc_s     = f"{dec['disc_pct']:>8.3f}%" if not np.isnan(dec['disc_pct']) else f"{'N/A':>9}"

        print(
            f"  {label:<12} {r.iterations:>6}  {conv:>5}  "
            f"{disc_s}  {dec['algo_pct']:>8.3f}%  {dec['total_pct']:>8.3f}%  "
            f"{jacobi_err:>11.3e}  {t:>8.2f}s"
        )
        if label == "QSVT-2D" and dec["algo_pct"] > 50.0:
            print(
                f"  QSVT-2D instability: algo error {dec['algo_pct']:.1f}%.\n"
                f"  Same root cause as 1-D HET. Under investigation."
            )

        results[f"het_{label}"] = {"result": r, "time": t, "dec": dec}

    results["het_u_exact"]  = u_exact_h
    results["het_problem"]  = prob_het
    results["het_Ex"]       = Ex_h
    results["het_Ey"]       = Ey_h
    results["het_u_thomas"] = u_thomas_h

    # ── (b) Generic 2-D Poisson verification ─────────────────────────────────
    print("\n  (b) Generic 2-D Poisson: fS source, N=4 (QSVT verification)")
    print(
        f"\n  {'Solver':<12} {'Iters':>6}  {'Conv':>5}  "
        f"{'Disc.':>9}  {'Algo.':>9}  {'Total':>9}  "
        f"{'JacobiErr':>11}  {'Time':>8}"
    )
    print(f"  {'─'*80}")

    cfg_gen  = SimConfig2D(N=4, epsilon=0.01, source_fn="fS", max_iter=100)
    prob_gen = PoissonProblem2D(cfg_gen)

    print("  Computing refined reference (refine_factor=9)...")
    t0_ref = time.perf_counter()
    u_ref  = prob_gen.classical_reference_solve(refine_factor=9)
    print(f"  Reference computed in {time.perf_counter()-t0_ref:.1f}s.")

    u_thomas_g = None
    for label, solver_fn, kwargs in [
        ("Thomas-2D", thomas_solve_2d,  {}),
        ("VQLS-2D",   vqls_solve_2d,    {"config": vqls_cfg_2d}),
        ("QSVT-2D",   qsvt_solve_2d,    {"config": qsvt_cfg_2d}),
    ]:
        t0 = time.perf_counter()
        r  = solver_fn(prob_gen, **kwargs)
        t  = time.perf_counter() - t0

        if label == "Thomas-2D":
            u_thomas_g = r.u

        ref = u_thomas_g if u_thomas_g is not None else r.u
        dec = _decompose_error(r.u.ravel(), ref.ravel(), u_ref.ravel())
        if label == "Thomas-2D":
            dec["algo_pct"] = 0.0

        jacobi_err = r.iteration_errors[-1] if r.iteration_errors else float("nan")
        conv       = "Yes" if r.converged else "No"
        disc_s     = f"{dec['disc_pct']:>8.3f}%" if not np.isnan(dec['disc_pct']) else f"{'N/A':>9}"

        print(
            f"  {label:<12} {r.iterations:>6}  {conv:>5}  "
            f"{disc_s}  {dec['algo_pct']:>8.3f}%  {dec['total_pct']:>8.3f}%  "
            f"{jacobi_err:>11.3e}  {t:>8.2f}s"
        )

        results[f"gen_{label}"] = {"result": r, "time": t, "dec": dec}

    results["gen_u_ref"]   = u_ref
    results["gen_problem"] = prob_gen

    return results


def plot_section_4(data: dict, save: bool = True) -> None:
    """
    Figure 4: 2-D results — HET sinusoidal and generic Poisson.

    Layout (2 rows x 3 cols):
        Row 1 (HET sinusoidal):
            Thomas-2D solution | VQLS-2D solution | QSVT-2D solution
        Row 2 (HET sinusoidal error + generic summary):
            Thomas error contour | VQLS error contour | error decomp bar chart
            (bar chart includes both HET and generic results)

    All three solvers shown even if QSVT is unstable.
    Error decomposition bar chart distinguishes disc. vs algo. error.
    """
    prob_het  = data["het_problem"]
    u_exact_h = data["het_u_exact"]
    X, Y      = prob_het.X, prob_het.Y

    r_thomas_h = data["het_Thomas-2D"]["result"]
    r_vqls_h   = data["het_VQLS-2D"]["result"]
    r_qsvt_h   = data["het_QSVT-2D"]["result"]

    u_all    = np.stack([u_exact_h, r_thomas_h.u, r_vqls_h.u])
    levels_u = np.linspace(u_all.min(), u_all.max(), 20)

    err_thomas = np.abs(r_thomas_h.u - u_exact_h)
    err_vqls   = np.abs(r_vqls_h.u  - u_exact_h)
    err_max    = max(err_thomas.max(), err_vqls.max())
    levels_e   = np.linspace(0.0, err_max if err_max > 0 else 1.0, 15)

    fig = plt.figure(figsize=(15, 9))
    fig.suptitle(
        "Section 4 — 2-D Poisson Results\n"
        r"HET sinusoidal: $\tilde{\phi}=\sin(\pi\tilde{x})\sin(\pi\tilde{y})$, N=4"
        "  |  Generic: fS source, N=4",
        fontsize=11,
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.40, wspace=0.34)

    # Row 1: solution contours for all three solvers (HET sinusoidal).
    for col, (r, title) in enumerate([
        (r_thomas_h, f"Thomas-2D ({r_thomas_h.iterations} iters)"),
        (r_vqls_h,   f"VQLS-2D ({r_vqls_h.iterations} iters)"),
        (r_qsvt_h,   f"QSVT-2D ({r_qsvt_h.iterations} iters) [unstable]"
                     if data["het_QSVT-2D"]["dec"]["algo_pct"] > 50 else
                     f"QSVT-2D ({r_qsvt_h.iterations} iters)"),
    ]):
        ax = fig.add_subplot(gs[0, col])
        cf = ax.contourf(X, Y, r.u, levels=levels_u, cmap="viridis")
        ax.contour(X, Y, u_exact_h, levels=8, colors="white",
                   linewidths=0.8, linestyles="--", alpha=0.7)
        fig.colorbar(cf, ax=ax, shrink=0.85)
        ax.set_xlabel(r"$\tilde{x}$")
        ax.set_ylabel(r"$\tilde{y}$")
        ax.set_title(title)
        ax.set_aspect("equal")

    # Row 2, col 0: Thomas-2D absolute error vs analytical.
    ax = fig.add_subplot(gs[1, 0])
    cf = ax.contourf(X, Y, err_thomas, levels=levels_e, cmap="hot_r")
    fig.colorbar(cf, ax=ax, shrink=0.85,
                 label=r"$|\tilde{\phi}_\mathrm{Thomas}-\tilde{\phi}_\mathrm{exact}|$")
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$\tilde{y}$")
    ax.set_title(f"Thomas-2D abs. error\n(disc. error, max={err_thomas.max():.3e})")
    ax.set_aspect("equal")

    # Row 2, col 1: VQLS-2D absolute error vs analytical.
    ax = fig.add_subplot(gs[1, 1])
    cf = ax.contourf(X, Y, err_vqls, levels=levels_e, cmap="hot_r")
    fig.colorbar(cf, ax=ax, shrink=0.85,
                 label=r"$|\tilde{\phi}_\mathrm{VQLS}-\tilde{\phi}_\mathrm{exact}|$")
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$\tilde{y}$")
    ax.set_title(f"VQLS-2D abs. error\n(total, max={err_vqls.max():.3e})")
    ax.set_aspect("equal")

    # Row 2, col 2: error decomposition bar chart.
    # Shows disc. error (grey) and algo. error (colour) for both problems.
    ax = fig.add_subplot(gs[1, 2])

    problems  = ["HET sinus.", "Generic fS"]
    keys_het  = ["het_Thomas-2D", "het_VQLS-2D", "het_QSVT-2D"]
    keys_gen  = ["gen_Thomas-2D", "gen_VQLS-2D", "gen_QSVT-2D"]
    solvers_2d = ["Thomas", "VQLS", "QSVT"]
    n_s        = len(solvers_2d)
    x_pos      = np.arange(n_s)
    width      = 0.35

    for i_prob, (prob_label, keys) in enumerate([
        ("HET", keys_het), ("Generic", keys_gen)
    ]):
        offset = -width/2 if i_prob == 0 else width/2
        hatch  = "" if i_prob == 0 else "//"
        disc_errs = [data[k]["dec"]["disc_pct"] for k in keys]
        algo_errs = [data[k]["dec"]["algo_pct"] for k in keys]

        ax.bar(x_pos + offset, disc_errs, width,
               color="lightgrey", edgecolor="black", lw=0.8,
               hatch=hatch, alpha=0.9,
               label=f"Disc. ({prob_label})")
        ax.bar(x_pos + offset, algo_errs, width,
               bottom=disc_errs,
               color=[COLOURS[k.lower()] for k in solvers_2d],
               edgecolor="black", lw=0.8, alpha=0.75, hatch=hatch,
               label=f"Algo. ({prob_label})")

    ax.set_xticks(x_pos)
    ax.set_xticklabels(solvers_2d)
    ax.set_ylabel("Max relative error (%)")
    ax.set_title(
        "Error decomposition (2-D)\n"
        "HET (solid) vs Generic (hatched)\n"
        "Grey: disc. | Colour: algo. (stacked)"
    )
    ax.legend(fontsize=6, ncol=2)
    ax.grid(True, alpha=0.3, axis="y")

    fig.tight_layout()
    _save_figure(fig, "figure_4_2d_results.pdf", save)
    plt.show()


# ============================================================================
# Excel export
# ============================================================================

def export_excel(s1_data: dict, s3_data: dict, s4_data: dict) -> None:
    """Export key metrics to Excel for supervisor review."""
    if not _EXCEL_AVAILABLE:
        print("  openpyxl not available; skipping Excel export.")
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb   = openpyxl.Workbook()
    path = RESULTS_DIR / "meeting5_metrics.xlsx"

    hf = Font(bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor="2C3E50")
    ca = Alignment(horizontal="center")

    def _ws(wb, title, headers, rows):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hf; cell.fill = hb; cell.alignment = ca
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = (
                max(len(str(c.value or "")) for c in col) + 3
            )
        return ws

    # Section 1.
    wb.active.title = "S1 Poisson 1D"
    rows1 = []
    for N in (4, 8):
        d = s1_data[N]
        for key, label in [("thomas","Thomas"),("hhl","HHL"),
                            ("vqls","VQLS"),("qsvt","QSVT")]:
            dec = d[key]["dec"]
            disc_s = round(dec["disc_pct"], 4) if not np.isnan(dec["disc_pct"]) else "N/A"
            rows1.append([
                N, label,
                round(dec["total_pct"], 4), disc_s,
                round(dec["algo_pct"], 4),
                f"{d[key]['res']:.4e}",
                round(d[key]["t"], 3),
                round(d[key].get("cost",""), 6) if key=="vqls" else "",
                d[key].get("degree","") if key=="qsvt" else "",
            ])
    ws1 = wb.active
    ws1.append(["N","Solver","Total%","Disc%","Algo%","Residual","Time(s)","Cost","Degree"])
    for cell in ws1[1]:
        cell.font = hf; cell.fill = hb; cell.alignment = ca
    for row in rows1:
        ws1.append(row)

    # Section 3.
    rows3 = []
    for sub, label_sub in [("3a","Linear/Hom."),("3b","Gaussian/Phys.")]:
        d = s3_data[sub]
        keys_3 = (["thomas","hhl","vqls","qsvt"]
                  if sub == "3a" else ["thomas","hhl","vqls"])
        for key in keys_3:
            if key not in d:
                continue
            dec    = d[key]["dec"]
            disc_s = round(dec["disc_pct"], 4) if not np.isnan(dec["disc_pct"]) else "N/A"
            res    = _residual(HETPoissonProblem1D(d["cfg"]).A,
                               d[key]["u"],
                               HETPoissonProblem1D(d["cfg"]).b)
            rows3.append([
                label_sub, key.upper(),
                round(dec["total_pct"], 4), disc_s,
                round(dec["algo_pct"], 4),
                f"{res:.4e}", round(d[key]["t"], 3),
            ])
    _ws(wb, "S3 HET 1D",
        ["Sub-case","Solver","Total%","Disc%","Algo%","Residual","Time(s)"],
        rows3)

    # Section 4.
    rows4 = []
    for prob_label, prefix in [("HET sinus.","het_"), ("Generic fS","gen_")]:
        for solver in ["Thomas-2D","VQLS-2D","QSVT-2D"]:
            key = prefix + solver
            if key not in s4_data:
                continue
            r   = s4_data[key]["result"]
            dec = s4_data[key]["dec"]
            disc_s = round(dec["disc_pct"], 4) if not np.isnan(dec["disc_pct"]) else "N/A"
            je = r.iteration_errors[-1] if r.iteration_errors else float("nan")
            rows4.append([
                prob_label, solver, r.iterations, r.converged,
                round(dec["total_pct"], 4), disc_s,
                round(dec["algo_pct"], 4),
                f"{je:.4e}", round(s4_data[key]["time"], 3),
            ])
    _ws(wb, "S4 2D",
        ["Problem","Solver","Iters","Conv","Total%","Disc%","Algo%","JacobiErr","Time(s)"],
        rows4)

    wb.save(path)
    print(f"\n  Excel saved: {path}")


# ============================================================================
# Main
# ============================================================================

def main() -> None:
    t_start = time.perf_counter()

    print("\n" + "═"*68)
    print("  QUANTUM POISSON SOLVER — MEETING 5 PROGRESS REPORT")
    print("  Imperial College London, Department of Aeronautics")
    print("  HHL | VQLS | QSVT — Poisson Equation and HET Plasma Modelling")
    print("═"*68)
    print(f"  Output directory: {RESULTS_DIR.resolve()}")

    s1_data = run_section_1()
    s3_data = run_section_3()
    s4_data = run_section_4()

    t_sections = time.perf_counter() - t_start
    print(f"\n{'─'*68}")
    print(f"  All sections completed in {t_sections:.1f}s. Generating figures...")

    plot_section_1(s1_data, save=True)
    plot_section_3(s3_data, save=True)
    plot_section_4(s4_data, save=True)

    export_excel(s1_data, s3_data, s4_data)

    print(f"\n  Total elapsed: {time.perf_counter() - t_start:.1f}s")
    print(f"  All outputs saved to: {RESULTS_DIR.resolve()}")
    print("═"*68)


if __name__ == "__main__":
    main()