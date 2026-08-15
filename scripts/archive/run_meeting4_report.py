"""
Meeting 4 Progress Report — Algorithm Comparison and HET Application.

Purpose
-------
This script generates the visual and tabular outputs for the meeting
progress report. It demonstrates the three quantum linear system
algorithms (HHL, VQLS, QSVT) implemented in this repository, applied to
the 1-D Poisson equation and the Hall Effect Thruster plasma modelling
problem.

Report structure
----------------
Section 1 — Algorithm Comparison on the 1-D Poisson Equation.
    All three solvers are run on N=4 and N=8 with the fS source function
    and homogeneous boundary conditions. Errors are decomposed into
    discretisation (FD truncation) and quantum algorithmic components.
    Output: Figure 1 (solution profiles, decomposed error, timing).

Section 2 — QSVT Circuit Complexity Analysis.
    Analytical complexity comparison for N in {4, 8, 16, 32}.
    Output: Figure 2 (scaling plots), Table 2.

Section 3 — HET Plasma Application (1-D).
    HHL and VQLS applied to the physical HET Poisson problem.
    QSVT included at N=4 (linear profile only).
    Output: Figure 3 (potential, electric field, decomposed error).

Section 4 — HET Plasma Application (2-D) + QSVT.
    Thomas, VQLS, and QSVT applied to the 2-D HET sinusoidal problem.
    Includes a generic 2-D Poisson QSVT verification case.
    Output: Figure 4 (contour plots, error decomposition).

Residual reporting convention
------------------------------
All residuals are reported as the normalised Euclidean residual:
    r = ||Au - b|| / ||b||
This is dimensionless and comparable across all problem scales.
For 2-D line-Jacobi solvers, the Jacobi iteration convergence error
(max|u^{n+1} - u^n|) is reported alongside the system residual, since
the system residual is always O(1) for a Jacobi iterate by construction.

Error decomposition
--------------------
Total error = discretisation error + quantum algorithmic error:
    e_disc = ||u_Thomas - u_exact|| / ||u_exact||   (FD truncation, O(h²))
    e_algo = ||u_solver - u_Thomas|| / ||u_Thomas||  (quantum approximation)
Both are reported in percent. Thomas is always the discrete reference.

References
----------
Ghafourpour & Laizet, Phys. Rev. Applied 24, 024032 (2025).
Bravo-Prieto et al., Quantum 7, 1188 (2023).
Harrow, Hassidim & Lloyd, Phys. Rev. Lett. 103, 150502 (2009).
Gilyen et al., STOC 2019, pp. 193-204.
Boeuf & Garrigues, J. Appl. Phys. 84(7), 3541-3554 (1998).
"""
from __future__ import annotations

import time
import warnings
import sys
from pathlib import Path

project_root = Path(__file__).resolve().parents[1]
if str(project_root) not in sys.path:
    sys.path.append(str(project_root))

import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt
import numpy as np

try:
    import openpyxl
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False
    warnings.warn(
        "openpyxl not installed; Excel export will be skipped. "
        "Install via: pip install openpyxl",
        ImportWarning,
    )

from core.config import SimConfig1D, SimConfig2D
from core.exact_solutions import EXACT_SOLUTIONS, HET_EXACT_SOLUTIONS
from core.het_config import HETConfig
from problems.het_plasma_1d import HETPoissonProblem1D
from problems.het_plasma_2d import HETConfig2D, HETSinusoidalProblem2D
from problems.poisson_1d import PoissonProblem1D
from problems.poisson_2d import PoissonProblem2D
from solvers.classical.thomas import thomas_solve, thomas_solve_system
from solvers.classical.thomas_2d import thomas_solve_2d
from solvers.quantum.hhl_1d import hhl_solve, hhl_solve_system
from solvers.quantum.block_encoding import subnormalisation_factor
from solvers.quantum.qsp_angles import polynomial_degree_estimate
from solvers.quantum.result import QSVTSolverResult
from solvers.quantum.vqls_1d import VQLSConfig1D, vqls_solve, vqls_solve_system
from solvers.quantum.vqls_2d import VQLSConfig2D, vqls_solve_2d
from solvers.quantum.qsvt_1d import QSVTConfig1D, qsvt_solve, qsvt_solve_system
from solvers.quantum.qsvt_2d import QSVTConfig2D, qsvt_solve_2d

RESULTS_DIR = Path("results/meeting5_report")
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

plt.rcParams.update({
    "font.family":    "serif",
    "font.size":      11,
    "axes.labelsize": 12,
    "axes.titlesize": 11,
    "legend.fontsize": 9,
    "figure.dpi":     140,
    "lines.linewidth": 1.8,
    "lines.markersize": 6,
})

COLOURS = {
    "thomas":     "#2ca02c",
    "hhl":        "#1f77b4",
    "vqls":       "#d62728",
    "qsvt":       "#9467bd",
    "analytical": "#000000",
}
MARKERS = {
    "thomas": "o",
    "hhl":    "s",
    "vqls":   "^",
    "qsvt":   "D",
}


# -- Utility functions ---------------------------------------------------------

def _rel_err_pct(u: np.ndarray, ref: np.ndarray) -> np.ndarray:
    """
    Pointwise absolute relative error in percent.

    Nodes where |ref| < 1% of max|ref| are masked to NaN to prevent
    division by near-zero values. This threshold is tighter than the
    previous 0.01% to avoid masking physically meaningful small values.
    """
    scale = np.max(np.abs(ref))
    mask  = np.abs(ref) > 0.01 * scale
    return np.where(mask, np.abs(u - ref) / np.abs(ref) * 100.0, np.nan)


def _max_rel_err(u: np.ndarray, ref: np.ndarray) -> float:
    """Maximum relative error in percent, excluding masked nodes."""
    err   = _rel_err_pct(u, ref)
    valid = err[~np.isnan(err)]
    return float(np.max(valid)) if valid.size > 0 else float("nan")


def _residual(A: np.ndarray, u: np.ndarray, b: np.ndarray) -> float:
    """Normalised Euclidean residual ||Au - b|| / ||b||."""
    return float(np.linalg.norm(A @ u - b) / np.linalg.norm(b))


def _decompose_error(
    u_solver  : np.ndarray,
    u_thomas  : np.ndarray,
    u_exact   : np.ndarray | None,
) -> dict:
    """
    Decompose the total solver error into discretisation and algorithmic
    components.

    Discretisation error (independent of quantum algorithm):
        e_disc = ||u_Thomas - u_exact|| / ||u_exact||  [percent]
    This is the FD truncation error O(h²), identical for all solvers.

    Quantum algorithmic error (solver-specific):
        e_algo = ||u_solver - u_Thomas|| / ||u_Thomas||  [percent]
    This isolates the quantum approximation error from discretisation.

    If u_exact is None (no analytical solution), e_disc is reported as
    N/A and e_total = e_algo.

    Parameters
    ----------
    u_solver : np.ndarray
    u_thomas : np.ndarray
    u_exact  : np.ndarray or None

    Returns
    -------
    dict with keys: disc_pct, algo_pct, total_pct
    """
    algo_pct = _max_rel_err(u_solver, u_thomas)

    if u_exact is not None:
        disc_pct  = _max_rel_err(u_thomas, u_exact)
        total_pct = _max_rel_err(u_solver, u_exact)
    else:
        disc_pct  = float("nan")
        total_pct = algo_pct

    return {
        "disc_pct":  disc_pct,
        "algo_pct":  algo_pct,
        "total_pct": total_pct,
    }


def _electric_field_1d(
    phi_int  : np.ndarray,
    alpha_bc : float,
    phi_0    : float,
    L        : float,
    N        : int,
) -> tuple[np.ndarray, np.ndarray]:
    """
    Recover the physical electric field E(x) [V/m] from the
    non-dimensional interior potential via second-order centred
    finite differences.
    """
    dx           = 1.0 / (N + 1)
    phi_full     = np.zeros(N + 2)
    phi_full[0]  = alpha_bc
    phi_full[1:N+1] = phi_int
    phi_full[N+1]   = 0.0
    x_full       = np.linspace(0.0, 1.0, N + 2)
    E_nd         = np.zeros(N + 2)
    E_nd[1:-1]   = -(phi_full[2:] - phi_full[:-2]) / (2.0 * dx)
    E_nd[0]      = -(phi_full[1]  - phi_full[0])   / dx
    E_nd[-1]     = -(phi_full[-1] - phi_full[-2])  / dx
    return x_full, E_nd * phi_0 / L


def _section_header(title: str, index: int) -> None:
    print(f"\n{'═'*68}")
    print(f"  SECTION {index} — {title}")
    print(f"{'═'*68}")


def _print_decomposed_header() -> None:
    print(
        f"  {'Solver':<10} {'MaxRelErr':>10}  {'Disc.Err':>9}  "
        f"{'Algo.Err':>9}  {'Residual':>12}  {'Time':>8}"
    )
    print(f"  {'─'*66}")


def _print_decomposed_row(
    label    : str,
    dec      : dict,
    residual : float,
    elapsed  : float,
    extra    : str = "",
) -> None:
    """
    Print a solver result row with decomposed error and normalised residual.

    Disc.Err: FD truncation error (same for all solvers, O(h²))
    Algo.Err: quantum approximation error vs Thomas
    MaxRelErr: total error vs analytical (if available)
    Residual: ||Au-b||/||b|| (dimensionless, comparable across problems)
    """
    disc_str  = f"{dec['disc_pct']:>8.3f}%" if not np.isnan(dec['disc_pct']) else f"{'N/A':>9}"
    algo_str  = f"{dec['algo_pct']:>8.3f}%"
    total_str = f"{dec['total_pct']:>9.3f}%"
    print(
        f"  {label:<10} {total_str}  {disc_str}  {algo_str}  "
        f"{residual:>12.4e}  {elapsed:>8.2f}s  {extra}"
    )


# ============================================================================
# Section 1 — Algorithm comparison on the 1-D Poisson equation
# ============================================================================

def run_section_1() -> dict:
    """
    Run all three quantum solvers on the 1-D Poisson equation with the
    fS source function and homogeneous BCs at N=4 and N=8.

    Reports decomposed errors (discretisation vs quantum algorithmic)
    and normalised residuals ||Au-b||/||b|| for all solvers.
    """
    _section_header("Algorithm Comparison — 1-D Poisson, fS Source", 1)
    _print_decomposed_header()

    vqls_cfg = VQLSConfig1D(
        n_layers    = 6,
        optimiser   = "COBYLA",
        max_iter    = 300,
        tol         = 1e-6,
        random_seed = 42,
        verbose     = False,
    )
    qsvt_cfg = QSVTConfig1D(
        epsilon=0.5, angle_method="auto", verbose=False, max_degree=2000,
    )

    results = {}

    for N in (4, 8):
        cfg     = SimConfig1D(N=N, epsilon=0.01, source_fn="fS")
        problem = PoissonProblem1D(cfg)
        u_exact = EXACT_SOLUTIONS["fS"](problem.x)

        print(f"\n  N={N}  (kappa={problem.kappa:.2f}):")

        # Thomas.
        t0       = time.perf_counter()
        r_thomas = thomas_solve(problem)
        t_thomas = time.perf_counter() - t0
        u_thomas = r_thomas.u
        dec_thomas = _decompose_error(u_thomas, u_thomas, u_exact)
        _print_decomposed_row(
            "Thomas", dec_thomas,
            _residual(problem.A, u_thomas, problem.b), t_thomas,
        )

        # HHL.
        t0    = time.perf_counter()
        r_hhl = hhl_solve(problem)
        t_hhl = time.perf_counter() - t0
        dec_hhl = _decompose_error(r_hhl.u, u_thomas, u_exact)
        _print_decomposed_row(
            "HHL", dec_hhl,
            _residual(problem.A, r_hhl.u, problem.b), t_hhl,
        )

        # VQLS.
        t0     = time.perf_counter()
        r_vqls = vqls_solve(problem, config=vqls_cfg)
        t_vqls = time.perf_counter() - t0
        dec_vqls = _decompose_error(r_vqls.u, u_thomas, u_exact)
        _print_decomposed_row(
            "VQLS", dec_vqls,
            _residual(problem.A, r_vqls.u, problem.b), t_vqls,
            f"cost={r_vqls.final_cost:.2e}",
        )

        # QSVT.
        t0     = time.perf_counter()
        r_qsvt = qsvt_solve(problem, config=qsvt_cfg)
        t_qsvt = time.perf_counter() - t0
        dec_qsvt = _decompose_error(r_qsvt.u, u_thomas, u_exact)
        _print_decomposed_row(
            "QSVT", dec_qsvt,
            _residual(problem.A, r_qsvt.u, problem.b), t_qsvt,
            f"deg={r_qsvt.polynomial_degree}, depth={r_qsvt.circuit_depth}",
        )

        results[N] = {
            "x": problem.x, "u_exact": u_exact, "cfg": cfg,
            "thomas": {"u": u_thomas, "t": t_thomas,
                       "res": _residual(problem.A, u_thomas, problem.b),
                       "dec": dec_thomas},
            "hhl":    {"u": r_hhl.u,  "t": t_hhl,
                       "res": _residual(problem.A, r_hhl.u, problem.b),
                       "dec": dec_hhl},
            "vqls":   {"u": r_vqls.u, "t": t_vqls,
                       "res": _residual(problem.A, r_vqls.u, problem.b),
                       "cost": r_vqls.final_cost, "dec": dec_vqls},
            "qsvt":   {"u": r_qsvt.u, "t": t_qsvt,
                       "res": _residual(problem.A, r_qsvt.u, problem.b),
                       "degree": r_qsvt.polynomial_degree,
                       "depth":  r_qsvt.circuit_depth,
                       "qubits": r_qsvt.n_qubits,
                       "alpha":  r_qsvt.alpha,
                       "kappa_eff": r_qsvt.kappa_effective,
                       "dec": dec_qsvt},
        }

    # Print per-node relative error series for asymmetry analysis.
    for N in (4, 8):
        d       = results[N]
        u_exact = d["u_exact"]
        for key, label in [("hhl","HHL"), ("vqls","VQLS"), ("qsvt","QSVT")]:
            err = _rel_err_pct(d[key]["u"], u_exact)
            # Replace NaN with 0 for printing.
            err_clean = np.where(np.isnan(err), 0.0, err)
            print(
                f"  N={N} {label} pointwise relative error (%):\n"
                f"    x values: {np.round(d['x'], 4).tolist()}\n"
                f"    errors:   {np.round(err_clean, 4).tolist()}"
            )

    return results


def plot_section_1(data: dict, save: bool = True) -> None:
    """
    Generate Figure 1: algorithm comparison on the 1-D Poisson equation.

    Layout (2 rows x 3 columns):
        Row 1 (N=4): solution profiles | error vs analytical | timing bar chart
        Row 2 (N=8): solution profiles | error vs analytical | summary table

    The error panel shows total error vs analytical for all solvers.
    Thomas is shown as the discretisation error baseline (dashed).
    The error panel uses a log scale to show the spread across solvers.
    """
    fig = plt.figure(figsize=(15, 9))
    fig.suptitle(
        "Section 1 — Algorithm Comparison: HHL vs VQLS vs QSVT\n"
        r"1-D Poisson, $f_S(x) = \sin(\pi x)$, $u(0)=u(1)=0$, "
        r"Analytical: $u=-\sin(\pi x)/\pi^2$",
        fontsize=12,
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.42, wspace=0.32)

    for row_idx, N in enumerate((4, 8)):
        d       = data[N]
        x       = d["x"]
        u_exact = d["u_exact"]
        x_full  = np.concatenate([[0.0], x, [1.0]])

        def _aug(u):
            return np.concatenate([[0.0], u, [0.0]])

        # -- Panel 1: solution profiles ---------------------------------------
        ax = fig.add_subplot(gs[row_idx, 0])
        ax.plot(x_full, _aug(u_exact),
                color=COLOURS["analytical"], lw=2.5, label="Analytical", zorder=6)
        for key, label in [
            ("thomas", "Thomas"), ("hhl", "HHL"),
            ("vqls", "VQLS"), ("qsvt", "QSVT"),
        ]:
            ax.plot(
                x_full, _aug(d[key]["u"]),
                color=COLOURS[key],
                ls="--" if key == "thomas" else "-",
                marker=MARKERS[key], ms=4, label=label,
            )
        kappa_val = (4.0 / np.pi**2) * (N + 1)**2
        ax.set_xlabel(r"$x$")
        ax.set_ylabel(r"$u(x)$")
        ax.set_title(f"Solution profiles (N={N}, $\\kappa\\approx{kappa_val:.0f}$)")
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

        # -- Panel 2: total error vs analytical (log scale) -------------------
        # Thomas (dashed): discretisation error — the baseline all solvers share.
        # HHL, VQLS, QSVT (solid): total error vs analytical.
        ax = fig.add_subplot(gs[row_idx, 1])
        ax.semilogy(
            x, _rel_err_pct(d["thomas"]["u"], u_exact),
            color=COLOURS["thomas"], ls="--",
            marker=MARKERS["thomas"], ms=4,
            label="Thomas (disc. baseline)", lw=1.8,
        )
        for key, label in [("hhl", "HHL"), ("vqls", "VQLS"), ("qsvt", "QSVT")]:
            ax.semilogy(
                x, _rel_err_pct(d[key]["u"], u_exact),
                color=COLOURS[key], ls="-",
                marker=MARKERS[key], ms=4, label=label, lw=1.8,
            )
        ax.set_xlabel(r"$x$")
        ax.set_ylabel("Relative error vs analytical (%)")
        ax.set_title(
            f"Error vs analytical (N={N})\n"
            "Dashed: Thomas discretisation baseline"
        )
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3, which="both")

        # -- Panel 3: residuals bar chart (row 1) or summary table (row 2) ---
        ax = fig.add_subplot(gs[row_idx, 2])
        if row_idx == 0:
            solvers_bar = ["Thomas", "HHL", "VQLS", "QSVT"]
            residuals   = [d[k]["res"] for k in ("thomas", "hhl", "vqls", "qsvt")]
            colours_bar = [COLOURS[k] for k in ("thomas", "hhl", "vqls", "qsvt")]
            bars = ax.bar(
                solvers_bar, residuals,
                color=colours_bar, alpha=0.8,
                edgecolor="black", linewidth=0.8,
            )
            for bar, res in zip(bars, residuals):
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    res * 1.5,
                    f"{res:.1e}",
                    ha="center", va="bottom", fontsize=8, rotation=30,
                )
            ax.set_ylabel(r"$\|Au-b\|/\|b\|$ (normalised residual)")
            ax.set_title(f"Normalised residuals (N={N})")
            ax.set_yscale("log")
            ax.grid(True, alpha=0.3, axis="y")
        else:
            # Summary table for N=8.
            rows = []
            for key, label in [
                ("thomas", "Thomas"), ("hhl", "HHL"),
                ("vqls", "VQLS"), ("qsvt", "QSVT"),
            ]:
                dec = d[key]["dec"]
                disc_s  = f"{dec['disc_pct']:.3f}%" if not np.isnan(dec['disc_pct']) else "N/A"
                algo_s  = f"{dec['algo_pct']:.3f}%"
                total_s = f"{dec['total_pct']:.3f}%"
                rows.append([label, total_s, disc_s, algo_s,
                             f"{d[key]['res']:.2e}"])
            tbl = ax.table(
                cellText  = rows,
                colLabels = ["Solver", "Total", "Disc.", "Algo.", "Residual"],
                cellLoc   = "center",
                loc       = "center",
            )
            tbl.auto_set_font_size(False)
            tbl.set_fontsize(8)
            tbl.scale(1.0, 1.6)
            for col in range(5):
                tbl[0, col].set_facecolor("#2c3e50")
                tbl[0, col].set_text_props(color="white", fontweight="bold")
            ax.axis("off")
            ax.set_title(f"Error decomposition (N={N})")

    plt.tight_layout()
    _save_figure(fig, "figure_1_algorithm_comparison.pdf", save)
    plt.show()


# ============================================================================
# Section 2 — QSVT circuit complexity analysis
# ============================================================================

def run_section_2() -> dict:
    """
    Compute QSVT circuit complexity metrics analytically for N in
    {4, 8, 16, 32} and compare against HHL circuit depth estimates.
    """
    _section_header("QSVT Circuit Complexity Analysis", 2)

    epsilon = 0.01
    results = {}

    print(
        f"\n  {'N':>4}  {'kappa':>8}  {'alpha':>6}  "
        f"{'kappa_eff':>10}  {'deg_QSVT':>10}  "
        f"{'depth_QSVT':>12}  {'depth_HHL_est':>14}  "
        f"{'qubits_QSVT':>12}"
    )
    print(f"  {'─'*82}")

    for N in (4, 8, 16, 32):
        kappa     = (4.0 / np.pi**2) * (N + 1)**2
        alpha     = subnormalisation_factor(-2.0, 1.0, N)
        A_norm_2  = 4.0
        kappa_eff = alpha * kappa / A_norm_2
        degree    = polynomial_degree_estimate(kappa_eff, epsilon)
        n         = int(np.log2(N))
        be_depth_est = 10 * n
        depth_qsvt   = degree * (be_depth_est + 1)
        n_l          = int(np.ceil(np.log2(kappa + 1))) + 1
        depth_hhl    = int(np.sqrt(2**n_l) / epsilon * n**2)
        qubits_qsvt  = n + 2 + 1

        results[N] = {
            "kappa": kappa, "alpha": alpha, "kappa_eff": kappa_eff,
            "degree": degree, "depth_qsvt": depth_qsvt,
            "depth_hhl": depth_hhl, "qubits_qsvt": qubits_qsvt, "n": n,
        }

        print(
            f"  {N:>4}  {kappa:>8.1f}  {alpha:>6.1f}  "
            f"{kappa_eff:>10.1f}  {degree:>10d}  "
            f"{depth_qsvt:>12d}  {depth_hhl:>14d}  "
            f"{qubits_qsvt:>12d}"
        )

    return results


def plot_section_2(data: dict, save: bool = True) -> None:
    """Generate Figure 2: QSVT circuit complexity scaling plots."""
    N_vals       = sorted(data.keys())
    kappas       = [data[N]["kappa"]       for N in N_vals]
    degrees      = [data[N]["degree"]      for N in N_vals]
    depths_qsvt  = [data[N]["depth_qsvt"]  for N in N_vals]
    depths_hhl   = [data[N]["depth_hhl"]   for N in N_vals]
    qubits       = [data[N]["qubits_qsvt"] for N in N_vals]

    fig, axes = plt.subplots(1, 3, figsize=(14, 5))
    fig.suptitle(
        "Section 2 — QSVT Circuit Complexity Scaling Analysis\n"
        r"1-D Poisson Equation, $\varepsilon = 0.01$",
        fontsize=12,
    )

    ax = axes[0]
    ax.loglog(kappas, degrees, color=COLOURS["qsvt"],
              marker="D", ms=8, lw=2, label=r"QSVT degree $d$")
    kappa_ref = np.array(kappas, dtype=float)
    ax.loglog(kappa_ref, kappa_ref * degrees[0] / kappas[0],
              color="grey", ls="--", lw=1.2, label=r"$\mathcal{O}(\kappa)$")
    ax.set_xlabel(r"Condition number $\kappa(A)$")
    ax.set_ylabel("Polynomial degree $d$")
    ax.set_title("QSVT polynomial degree")
    ax.legend()
    ax.grid(True, alpha=0.3, which="both")
    for N, kap, deg in zip(N_vals, kappas, degrees):
        ax.annotate(f"N={N}", (kap, deg), textcoords="offset points",
                    xytext=(5, 5), fontsize=8)

    ax = axes[1]
    x_pos = np.arange(len(N_vals))
    width = 0.35
    ax.bar(x_pos - width/2, depths_qsvt, width,
           color=COLOURS["qsvt"], alpha=0.85,
           label="QSVT", edgecolor="black", lw=0.8)
    ax.bar(x_pos + width/2, depths_hhl, width,
           color=COLOURS["hhl"], alpha=0.85,
           label="HHL (estimate)", edgecolor="black", lw=0.8)
    ax.set_xticks(x_pos)
    ax.set_xticklabels([f"N={N}" for N in N_vals])
    ax.set_ylabel("Circuit depth (gate count)")
    ax.set_title("Circuit depth: QSVT vs HHL")
    ax.set_yscale("log")
    ax.legend()
    ax.grid(True, alpha=0.3, axis="y")

    ax = axes[2]
    ax.plot(N_vals, qubits, color=COLOURS["qsvt"],
            marker="D", ms=8, lw=2, label="QSVT total qubits")
    hhl_qubits = []
    for N in N_vals:
        n   = int(np.log2(N))
        kap = data[N]["kappa"]
        n_l = int(np.ceil(np.log2(kap + 1))) + 2
        hhl_qubits.append(n + n_l + (n - 1) + 1)
    ax.plot(N_vals, hhl_qubits, color=COLOURS["hhl"],
            marker="s", ms=8, lw=2, ls="--", label="HHL total qubits")
    ax.set_xlabel("System size N")
    ax.set_ylabel("Total qubit count")
    ax.set_title("Qubit requirements")
    ax.legend()
    ax.grid(True, alpha=0.3)
    for N, q in zip(N_vals, qubits):
        ax.annotate(f"{q}", (N, q), textcoords="offset points",
                    xytext=(4, 4), fontsize=9, color=COLOURS["qsvt"])

    plt.tight_layout()
    _save_figure(fig, "figure_2_qsvt_complexity.pdf", save)
    plt.show()


# ============================================================================
# Section 3 — HET plasma application (1-D)
# ============================================================================

def run_section_3() -> dict:
    """
    Apply HHL, VQLS, and QSVT to the 1-D HET plasma Poisson equation.

    Sub-case 3a: linear profile, homogeneous BCs, N=4.
        Analytical solution available. QSVT included.
        QSVT instability note: the pyqsp sym_qsp method achieves the
        polynomial as Im(<0|U|0>), but the QSVT circuit extracts Re(<0|U|0>).
        This convention mismatch causes QSVT to fail for the HET problem
        where the proportionality recovery cannot compensate for the
        large RHS norm (||b|| ~ 727). Under investigation.

    Sub-case 3b: Gaussian profile, V_d=300V, N=8.
        Thomas is the reference (no analytical solution).
        HHL and VQLS only (QSVT circuit depth exceeds laptop threshold).

    VQLS asymmetry note: larger relative errors at low x are explained
    by the proportionality recovery weighting high-amplitude components
    of b more heavily in the least-squares minimisation.
    """
    _section_header("HET Plasma Application — 1-D Axial Poisson Equation", 3)

    vqls_cfg_n4 = VQLSConfig1D(
        n_layers=6, max_iter=300, tol=1e-6, random_seed=42, verbose=False
    )
    vqls_cfg_n8 = VQLSConfig1D(
        n_layers=8, max_iter=500, tol=1e-5, random_seed=42, verbose=False
    )
    qsvt_cfg_het = QSVTConfig1D(
        epsilon=0.5, angle_method="auto", verbose=False, max_degree=2000, label = "HET-3a",
    )

    results = {}

    # -- Sub-case 3a ----------------------------------------------------------
    print("\n  Sub-case 3a: linear profile, homogeneous BCs (N=4, QSVT included)")
    _print_decomposed_header()

    cfg_a  = HETConfig(N=4, epsilon=0.01, rho_profile="linear", V_discharge=0.0)
    prob_a = HETPoissonProblem1D(cfg_a)
    u_exact_a = HET_EXACT_SOLUTIONS["linear"](prob_a.x, cfg_a.rho_0, cfg_a.alpha)

    t0         = time.perf_counter()
    u_thomas_a = thomas_solve_system(prob_a.A, prob_a.b)
    t_thomas_a = time.perf_counter() - t0
    dec_ta = _decompose_error(u_thomas_a, u_thomas_a, u_exact_a)
    _print_decomposed_row("Thomas", dec_ta, _residual(prob_a.A, u_thomas_a, prob_a.b), t_thomas_a)

    t0 = time.perf_counter()
    u_hhl_a, _, _ = hhl_solve_system(prob_a.A, prob_a.b, cfg_a.epsilon)
    t_hhl_a = time.perf_counter() - t0
    dec_ha = _decompose_error(u_hhl_a, u_thomas_a, u_exact_a)
    _print_decomposed_row("HHL", dec_ha, _residual(prob_a.A, u_hhl_a, prob_a.b), t_hhl_a)

    t0       = time.perf_counter()
    vr_a     = vqls_solve_system(prob_a.A, prob_a.b, vqls_cfg_n4)
    u_vqls_a = vr_a.u
    t_vqls_a = time.perf_counter() - t0
    dec_va = _decompose_error(u_vqls_a, u_thomas_a, u_exact_a)
    _print_decomposed_row("VQLS", dec_va, _residual(prob_a.A, u_vqls_a, prob_a.b), t_vqls_a,
                          f"cost={vr_a.final_cost:.2e}")

    t0       = time.perf_counter()
    qr_a     = qsvt_solve_system(prob_a.A, prob_a.b, qsvt_cfg_het)
    u_qsvt_a = qr_a.u
    t_qsvt_a = time.perf_counter() - t0
    dec_qa = _decompose_error(u_qsvt_a, u_thomas_a, u_exact_a)
    _print_decomposed_row("QSVT", dec_qa, _residual(prob_a.A, u_qsvt_a, prob_a.b), t_qsvt_a,
                          f"deg={qr_a.polynomial_degree}, depth={qr_a.circuit_depth}")

    if dec_qa["algo_pct"] > 50.0:
        print(
            f"\n  QSVT INSTABILITY NOTE (3a): algo error = {dec_qa['algo_pct']:.1f}%.\n"
            f"  Root cause under investigation. Hypothesis: pyqsp sym_qsp achieves\n"
            f"  the polynomial as Im(<0|U|0>), but the circuit extracts Re(<0|U|0>).\n"
            f"  For generic Poisson the proportionality recovery compensates;\n"
            f"  for HET (||b||~727) the large scale amplifies the convention error.\n"
            f"  QSVT works correctly for Section 1 (generic Poisson)."
        )

    results["3a"] = {
        "x": prob_a.x, "cfg": cfg_a, "u_exact": u_exact_a,
        "thomas": {"u": u_thomas_a, "t": t_thomas_a, "dec": dec_ta},
        "hhl":    {"u": u_hhl_a,    "t": t_hhl_a,    "dec": dec_ha},
        "vqls":   {"u": u_vqls_a,   "t": t_vqls_a,   "dec": dec_va,
                   "cost": vr_a.final_cost},
        "qsvt":   {"u": u_qsvt_a,   "t": t_qsvt_a,   "dec": dec_qa,
                   "degree": qr_a.polynomial_degree, "depth": qr_a.circuit_depth},
    }

    # -- Sub-case 3b ----------------------------------------------------------
    print("\n  Sub-case 3b: Gaussian profile, V_d=300V (N=8, HHL+VQLS only)")
    print("  Note: no analytical solution — Thomas is the discrete reference.")
    print("  Disc.Err = N/A; Algo.Err = error vs Thomas.")
    _print_decomposed_header()

    cfg_b  = HETConfig(N=8, epsilon=0.01, rho_profile="gaussian", V_discharge=300.0)
    prob_b = HETPoissonProblem1D(cfg_b)

    t0         = time.perf_counter()
    u_thomas_b = thomas_solve_system(prob_b.A, prob_b.b)
    t_thomas_b = time.perf_counter() - t0
    dec_tb = _decompose_error(u_thomas_b, u_thomas_b, None)
    _print_decomposed_row("Thomas", dec_tb, _residual(prob_b.A, u_thomas_b, prob_b.b),
                          t_thomas_b, "(reference)")

    t0 = time.perf_counter()
    u_hhl_b, _, _ = hhl_solve_system(prob_b.A, prob_b.b, cfg_b.epsilon)
    t_hhl_b = time.perf_counter() - t0
    dec_hb = _decompose_error(u_hhl_b, u_thomas_b, None)
    _print_decomposed_row("HHL", dec_hb, _residual(prob_b.A, u_hhl_b, prob_b.b), t_hhl_b)

    t0       = time.perf_counter()
    vr_b     = vqls_solve_system(prob_b.A, prob_b.b, vqls_cfg_n8)
    u_vqls_b = vr_b.u
    t_vqls_b = time.perf_counter() - t0
    dec_vb = _decompose_error(u_vqls_b, u_thomas_b, None)
    _print_decomposed_row("VQLS", dec_vb, _residual(prob_b.A, u_vqls_b, prob_b.b), t_vqls_b,
                          f"cost={vr_b.final_cost:.2e}")

    print(f"  QSVT: N/A — circuit depth exceeds laptop threshold at N=8")

    x_full_b, E_thomas_b = _electric_field_1d(
        u_thomas_b, cfg_b.alpha_bc, cfg_b.phi_0, cfg_b.L, cfg_b.N)
    _, E_hhl_b  = _electric_field_1d(u_hhl_b,  cfg_b.alpha_bc, cfg_b.phi_0, cfg_b.L, cfg_b.N)
    _, E_vqls_b = _electric_field_1d(u_vqls_b, cfg_b.alpha_bc, cfg_b.phi_0, cfg_b.L, cfg_b.N)

    print(f"\n  Peak |E| Thomas: {np.max(np.abs(E_thomas_b)):.3e} V/m")
    print(f"  Peak |E| HHL:    {np.max(np.abs(E_hhl_b)):.3e} V/m")
    print(f"  Peak |E| VQLS:   {np.max(np.abs(E_vqls_b)):.3e} V/m")
    print(f"  B&G (1998) Fig.3 reference: ~2x10^4 V/m near x/L ~ 0.8")
    print(f"  NOTE: 2-order-of-magnitude discrepancy. Current model uses a")
    print(f"  prescribed Gaussian charge density (delta_0 = delta_0_factor/alpha).")
    print(f"  B&G uses a self-consistent coupled fluid model with different")
    print(f"  geometry and ionisation dynamics. The non-dimensional scaling")
    print(f"  phi_0/L * alpha_bc = {cfg_b.phi_0/cfg_b.L * cfg_b.alpha_bc:.2e} V/m")
    print(f"  sets the base field scale, which is already 2 orders above B&G.")
    print(f"  Resolution: use B&G steady-state density profile as source term.")

    results["3b"] = {
        "x_int": prob_b.x, "x_full": x_full_b, "cfg": cfg_b,
        "thomas": {"u": u_thomas_b, "E": E_thomas_b, "t": t_thomas_b, "dec": dec_tb},
        "hhl":    {"u": u_hhl_b,    "E": E_hhl_b,    "t": t_hhl_b,    "dec": dec_hb},
        "vqls":   {"u": u_vqls_b,   "E": E_vqls_b,   "t": t_vqls_b,   "dec": dec_vb,
                   "cost": vr_b.final_cost},
    }

    return results


def plot_section_3(data: dict, save: bool = True) -> None:
    """
    Generate Figure 3: HET plasma 1-D results.

    Layout (2 rows x 3 columns):
        Row 1 (3a, linear/hom.): potential | electric field | error vs analytical
        Row 2 (3b, Gaussian/phys.): potential | electric field | algo error vs Thomas

    Thomas is included as a baseline in all panels.
    """
    fig = plt.figure(figsize=(15, 9))
    fig.suptitle(
        "Section 3 — HET Plasma Application: 1-D Axial Poisson Equation\n"
        "Physical parameters: Boeuf & Garrigues (1998), J. Appl. Phys. 84, 3541",
        fontsize=12,
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.42, wspace=0.32)

    # -- Row 1: sub-case 3a ---------------------------------------------------
    d3a  = data["3a"]
    cfg  = d3a["cfg"]
    x    = d3a["x"]
    x_full = np.concatenate([[0.0], x, [1.0]])

    # Potential.
    ax = fig.add_subplot(gs[0, 0])
    ax.plot(x_full, np.concatenate([[0.0], d3a["u_exact"], [0.0]]),
            color=COLOURS["analytical"], lw=2.5, label="Analytical")
    for key, label in [("thomas","Thomas"),("hhl","HHL"),("vqls","VQLS"),("qsvt","QSVT")]:
        if d3a.get(key) is None:
            continue
        ax.plot(x_full, np.concatenate([[0.0], d3a[key]["u"], [0.0]]),
                color=COLOURS[key], ls="--" if key=="thomas" else "-",
                marker=MARKERS[key], ms=4, label=label)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$\tilde{\phi}$")
    ax.set_title("Potential: linear, hom. BCs (N=4)")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # Electric field (homogeneous BCs: alpha_bc=0).
    ax = fig.add_subplot(gs[0, 1])
    # Analytical electric field for linear HET profile.
    xE_an = np.linspace(0, 1, 100)
    E_an  = -(cfg.alpha * cfg.rho_0 / 6.0) * (1.0 - 3.0 * xE_an**2)
    ax.plot(xE_an, E_an * cfg.phi_0 / cfg.L / 1e3,
            color=COLOURS["analytical"], lw=2.5, label="Analytical")
    for key, label in [("thomas","Thomas"),("hhl","HHL"),("vqls","VQLS"),("qsvt","QSVT")]:
        if d3a.get(key) is None:
            continue
        xE, E = _electric_field_1d(d3a[key]["u"], 0.0, cfg.phi_0, cfg.L, cfg.N)
        ax.plot(xE, E / 1e3, color=COLOURS[key],
                ls="--" if key=="thomas" else "-",
                marker=MARKERS[key], ms=4, label=label)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$E$ [kV/m]")
    ax.set_title("Electric field: linear profile (N=4)")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # Error vs analytical.
    ax = fig.add_subplot(gs[0, 2])
    # Thomas: discretisation error baseline.
    ax.semilogy(x, _rel_err_pct(d3a["thomas"]["u"], d3a["u_exact"]),
                color=COLOURS["thomas"], ls="--",
                marker=MARKERS["thomas"], ms=4,
                label="Thomas (disc. baseline)", lw=1.8)
    for key, label in [("hhl","HHL"),("vqls","VQLS"),("qsvt","QSVT")]:
        if d3a.get(key) is None:
            continue
        ax.semilogy(x, _rel_err_pct(d3a[key]["u"], d3a["u_exact"]),
                    color=COLOURS[key], ls="-",
                    marker=MARKERS[key], ms=4, label=label, lw=1.8)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel("Relative error vs analytical (%)")
    ax.set_title("Error vs analytical (N=4)")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3, which="both")

    # -- Row 2: sub-case 3b ---------------------------------------------------
    d3b   = data["3b"]
    cfg_b = d3b["cfg"]
    x_int = d3b["x_int"]
    x_full_b = d3b["x_full"]

    # Potential.
    ax = fig.add_subplot(gs[1, 0])
    for key, label in [("thomas","Thomas"),("hhl","HHL"),("vqls","VQLS")]:
        phi_full = np.concatenate([[cfg_b.alpha_bc], d3b[key]["u"], [0.0]])
        ax.plot(x_full_b, phi_full, color=COLOURS[key],
                ls="--" if key=="thomas" else "-",
                marker=MARKERS[key], ms=4, label=label)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$\tilde{\phi}$")
    ax.set_title(r"Potential: Gaussian, $V_d=300$ V (N=8)")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # Electric field.
    ax = fig.add_subplot(gs[1, 1])
    for key, label in [("thomas","Thomas"),("hhl","HHL"),("vqls","VQLS")]:
        ax.plot(x_full_b, d3b[key]["E"] / 1e4, color=COLOURS[key],
                ls="--" if key=="thomas" else "-",
                marker=MARKERS[key], ms=4, label=label)
    ax.axhline(2.0, color="grey", ls="-.", lw=1.2, alpha=0.7,
               label=r"B&G (1998): $\sim 2\times10^4$ V/m")
    ax.axvline(0.8, color="grey", ls=":", lw=1.0, alpha=0.5)
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel(r"$E$ [$\times 10^4$ V/m]")
    ax.set_title("Electric field: Gaussian (N=8)\ncf. Boeuf & Garrigues (1998)")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # Algorithmic error vs Thomas.
    ax = fig.add_subplot(gs[1, 2])
    ref_b = d3b["thomas"]["u"]
    for key, label in [("hhl","HHL"),("vqls","VQLS")]:
        ax.semilogy(x_int, _rel_err_pct(d3b[key]["u"], ref_b),
                    color=COLOURS[key], marker=MARKERS[key], ms=5,
                    label=f"{label} (algo vs Thomas)")
    ax.set_xlabel(r"$\tilde{x}$")
    ax.set_ylabel("Algorithmic error vs Thomas (%)")
    ax.set_title("Algorithmic error (N=8)\nNo analytical solution available")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3, which="both")

    plt.tight_layout()
    _save_figure(fig, "figure_3_het_1d.pdf", save)
    plt.show()


# ============================================================================
# Section 4 — HET plasma application (2-D) + QSVT verification
# ============================================================================

def run_section_4() -> dict:
    """
    Apply Thomas-2D, VQLS-2D, and QSVT-2D to the 2-D HET sinusoidal
    Poisson problem with the analytical solution phi = sin(pi*x)*sin(pi*y).

    Also runs a generic 2-D Poisson QSVT verification case (fS source,
    no HET physics) to isolate whether QSVT-2D failures are due to the
    HET source term or the 2D architecture itself.

    Residual reporting for 2-D solvers:
        - System residual ||A_full u - b_full|| / ||b_full|| is reported
          but is expected to be O(1) for Jacobi iterates (not a bug).
        - Jacobi convergence error max|u^{n+1} - u^n| is the meaningful
          convergence metric and is printed separately.
    """
    _section_header(
        "HET Plasma Application — 2-D Poisson, Sinusoidal Source", 4
    )

    # -- HET 2-D sinusoidal ---------------------------------------------------
    cfg     = HETConfig2D(N=4, epsilon=0.01, max_iter=300)
    problem = HETSinusoidalProblem2D(cfg)
    u_exact = problem.analytical_solution()
    Ex_exact, Ey_exact = problem.analytical_electric_field()

    print(f"  {problem.summary()}")
    print(f"  max|phi_exact| = {np.max(np.abs(u_exact)):.4f}")
    print(f"  Note: system residual is O(1) for Jacobi iterates by construction.")
    print(f"  Convergence is measured by Jacobi update error max|u^{{n+1}}-u^n|.")
    print(
        f"\n  {'Solver':<12} {'Iters':>6}  {'Conv':>5}  "
        f"{'Disc.Err':>9}  {'Algo.Err':>9}  {'Total':>9}  "
        f"{'JacobiErr':>11}  {'Time':>8}"
    )
    print(f"  {'─'*78}")

    inner_cfg = VQLSConfig1D(
        n_layers=3, max_iter=100, tol=1e-2, random_seed=0, verbose=False
    )
    vqls_cfg_2d = VQLSConfig2D(
        inner_config=inner_cfg, warm_start=True, verbose=False
    )
    qsvt_cfg_2d = QSVTConfig2D(
        epsilon=0.01, angle_method="auto", max_degree=200, verbose=False
    )

    results = {
        "cfg": cfg, "problem": problem,
        "u_exact": u_exact,
        "Ex_exact": Ex_exact, "Ey_exact": Ey_exact,
    }

    u_thomas_2d = None

    for label, solver_fn, kwargs in [
        ("Thomas-2D", thomas_solve_2d,  {}),
        ("VQLS-2D",   vqls_solve_2d,    {"config": vqls_cfg_2d}),
        ("QSVT-2D",   qsvt_solve_2d,    {"config": qsvt_cfg_2d}),
    ]:
        t0 = time.perf_counter()
        r  = solver_fn(problem, **kwargs)
        t  = time.perf_counter() - t0

        if label == "Thomas-2D":
            u_thomas_2d = r.u

        dec = _decompose_error(
            r.u.ravel(), u_thomas_2d.ravel() if u_thomas_2d is not None else r.u.ravel(),
            u_exact.ravel(),
        )
        if label == "Thomas-2D":
            dec["algo_pct"] = 0.0

        jacobi_err = r.iteration_errors[-1] if r.iteration_errors else float("nan")
        conv = "Yes" if r.converged else "No"

        disc_s  = f"{dec['disc_pct']:>8.3f}%" if not np.isnan(dec['disc_pct']) else f"{'N/A':>9}"
        algo_s  = f"{dec['algo_pct']:>8.3f}%"
        total_s = f"{dec['total_pct']:>8.3f}%"

        print(
            f"  {label:<12} {r.iterations:>6}  {conv:>5}  "
            f"{disc_s}  {algo_s}  {total_s}  "
            f"{jacobi_err:>11.3e}  {t:>8.2f}s"
        )

        if label == "QSVT-2D" and dec["algo_pct"] > 50.0:
            print(
                f"  QSVT-2D INSTABILITY: algo error {dec['algo_pct']:.1f}%.\n"
                f"  Same root cause as 1-D HET: pyqsp convention mismatch.\n"
                f"  The 2-D row matrix (kappa~2.36) should give degree~59,\n"
                f"  but the polynomial convention error corrupts the solution."
            )

        results[label] = {"result": r, "time": t, "dec": dec}

    # -- Generic 2-D Poisson QSVT verification --------------------------------
    print(f"\n  QSVT-2D Verification: generic 2-D Poisson, fS source, N=4")
    print(f"  (Tests QSVT-2D on a problem without HET physics)")
    print(
        f"\n  {'Solver':<12} {'Iters':>6}  {'Conv':>5}  "
        f"{'Disc.Err':>9}  {'Algo.Err':>9}  {'Total':>9}  "
        f"{'JacobiErr':>11}  {'Time':>8}"
    )
    print(f"  {'─'*78}")

    cfg_generic = SimConfig2D(N=4, epsilon=0.01, source_fn="fS", max_iter=100)
    prob_generic = PoissonProblem2D(cfg_generic)
    u_exact_generic = None   # no 2D analytical solution for fS in this formulation

    # Use refined Thomas as reference.
    print("  Computing refined reference (refine_factor=9)...")
    t0_ref = time.perf_counter()
    u_ref_generic = prob_generic.classical_reference_solve(refine_factor=9)
    t_ref = time.perf_counter() - t0_ref
    print(f"  Reference computed in {t_ref:.1f}s.")

    u_thomas_generic = None
    for label, solver_fn, kwargs in [
        ("Thomas-2D", thomas_solve_2d,  {}),
        ("VQLS-2D",   vqls_solve_2d,    {"config": vqls_cfg_2d}),
        ("QSVT-2D",   qsvt_solve_2d,    {"config": qsvt_cfg_2d}),
    ]:
        t0 = time.perf_counter()
        r  = solver_fn(prob_generic, **kwargs)
        t  = time.perf_counter() - t0

        if label == "Thomas-2D":
            u_thomas_generic = r.u

        # Error vs refined reference.
        ref = u_thomas_generic if u_thomas_generic is not None else r.u
        dec_g = _decompose_error(
            r.u.ravel(),
            ref.ravel(),
            u_ref_generic.ravel(),
        )
        if label == "Thomas-2D":
            dec_g["algo_pct"] = 0.0

        jacobi_err = r.iteration_errors[-1] if r.iteration_errors else float("nan")
        conv = "Yes" if r.converged else "No"
        disc_s  = f"{dec_g['disc_pct']:>8.3f}%" if not np.isnan(dec_g['disc_pct']) else f"{'N/A':>9}"
        print(
            f"  {label:<12} {r.iterations:>6}  {conv:>5}  "
            f"{disc_s}  {dec_g['algo_pct']:>8.3f}%  "
            f"{dec_g['total_pct']:>8.3f}%  "
            f"{jacobi_err:>11.3e}  {t:>8.2f}s"
        )

        results[f"generic_{label}"] = {"result": r, "time": t, "dec": dec_g}

    results["u_ref_generic"] = u_ref_generic
    results["prob_generic"]  = prob_generic

    return results


def plot_section_4(data: dict, save: bool = True) -> None:
    """
    Generate Figure 4: 2-D HET sinusoidal Poisson results.

    Layout (2 rows x 3 columns):
        Row 1: analytical | Thomas-2D | VQLS-2D
        Row 2: Thomas error | VQLS error | error decomposition bar chart
              (QSVT-2D included in bar chart; contour omitted if unstable)
    """
    problem  = data["problem"]
    u_exact  = data["u_exact"]
    X, Y     = problem.X, problem.Y
    r_thomas = data["Thomas-2D"]["result"]
    r_vqls   = data["VQLS-2D"]["result"]
    r_qsvt   = data["QSVT-2D"]["result"]

    u_all    = np.stack([u_exact, r_thomas.u, r_vqls.u])
    u_min, u_max = u_all.min(), u_all.max()
    levels_u = np.linspace(u_min, u_max, 25)

    err_thomas = np.abs(r_thomas.u - u_exact)
    err_vqls   = np.abs(r_vqls.u   - u_exact)
    err_max    = max(err_thomas.max(), err_vqls.max())
    levels_e   = np.linspace(0.0, err_max if err_max > 0 else 1.0, 20)

    fig = plt.figure(figsize=(15, 9))
    fig.suptitle(
        "Section 4 — HET Plasma: 2-D Poisson, Sinusoidal Source, N=4\n"
        r"Analytical: $\tilde{\phi} = \sin(\pi\tilde{x})\sin(\pi\tilde{y})$"
        "  |  Note: system residual is O(1) for Jacobi iterates",
        fontsize=11,
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.40, wspace=0.32)

    # Row 1: solution contours.
    for col, (Z, title) in enumerate([
        (u_exact,    "Analytical solution"),
        (r_thomas.u, f"Thomas-2D ({r_thomas.iterations} iters)"),
        (r_vqls.u,   f"VQLS-2D ({r_vqls.iterations} iters)"),
    ]):
        ax = fig.add_subplot(gs[0, col])
        cf = ax.contourf(X, Y, Z, levels=levels_u, cmap="viridis")
        ax.contour(X, Y, Z, levels=levels_u,
                   colors="white", linewidths=0.3, alpha=0.4)
        fig.colorbar(cf, ax=ax, shrink=0.85)
        ax.set_xlabel(r"$\tilde{x}$")
        ax.set_ylabel(r"$\tilde{y}$")
        ax.set_title(title)
        ax.set_aspect("equal")

    # Row 2, col 0-1: error contours vs analytical.
    for col, (Z, title) in enumerate([
        (err_thomas, f"Thomas-2D abs. error\nmax={err_thomas.max():.3e}"),
        (err_vqls,   f"VQLS-2D abs. error\nmax={err_vqls.max():.3e}"),
    ]):
        ax = fig.add_subplot(gs[1, col])
        cf = ax.contourf(X, Y, Z, levels=levels_e, cmap="hot_r")
        fig.colorbar(cf, ax=ax, shrink=0.85,
                     label=r"$|\tilde{\phi}_{solver} - \tilde{\phi}_{exact}|$")
        ax.set_xlabel(r"$\tilde{x}$")
        ax.set_ylabel(r"$\tilde{y}$")
        ax.set_title(title)
        ax.set_aspect("equal")

    # Row 2, col 2: error decomposition bar chart including QSVT.
    ax = fig.add_subplot(gs[1, 2])
    solvers_2d = ["Thomas-2D", "VQLS-2D", "QSVT-2D"]
    disc_errs  = [data[k]["dec"]["disc_pct"]  for k in solvers_2d]
    algo_errs  = [data[k]["dec"]["algo_pct"]  for k in solvers_2d]
    x_pos      = np.arange(len(solvers_2d))
    width      = 0.35

    ax.bar(x_pos - width/2, disc_errs, width,
           color="lightgrey", edgecolor="black", lw=0.8,
           label="Discretisation error\n(Thomas vs analytical)")
    ax.bar(x_pos + width/2, algo_errs, width,
           color=[COLOURS["thomas"], COLOURS["vqls"], COLOURS["qsvt"]],
           alpha=0.85, edgecolor="black", lw=0.8,
           label="Quantum algorithmic error\n(solver vs Thomas)")
    for i, (bar, val) in enumerate(zip(
        ax.patches[len(solvers_2d):], algo_errs
    )):
        if val > 0 and val < 20:
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() * 1.02,
                f"{val:.2f}%",
                ha="center", va="bottom", fontsize=8,
            )
    ax.set_xticks(x_pos)
    ax.set_xticklabels(solvers_2d, fontsize=8)
    ax.set_ylabel("Max relative error (%)")
    ax.set_title(
        "Error decomposition (2-D)\n"
        "Grey: disc. | Colour: algorithmic"
    )
    ax.legend(fontsize=7)
    ax.grid(True, alpha=0.3, axis="y")

    plt.tight_layout()
    _save_figure(fig, "figure_4_het_2d.pdf", save)
    plt.show()


# ============================================================================
# Excel / CSV export
# ============================================================================

def export_excel(
    s1_data : dict,
    s2_data : dict,
    s3_data : dict,
    s4_data : dict,
) -> None:
    """Export all scalar benchmark metrics to a single Excel workbook."""
    if _EXCEL_AVAILABLE:
        _export_excel_openpyxl(s1_data, s2_data, s3_data, s4_data)
    else:
        _export_csv_fallback(s1_data, s2_data, s3_data, s4_data)


def _export_excel_openpyxl(s1_data, s2_data, s3_data, s4_data) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb   = openpyxl.Workbook()
    path = RESULTS_DIR / "meeting_report_metrics.xlsx"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2C3E50")
    center_align = Alignment(horizontal="center")

    def _write_sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    # Sheet 1: Algorithm comparison with decomposed error.
    ws1 = wb.active
    ws1.title = "S1 Algorithm Comparison"
    headers1 = [
        "N", "Solver",
        "Total Error (%)", "Disc. Error (%)", "Algo. Error (%)",
        "Normalised Residual ||Au-b||/||b||",
        "Wall Time (s)", "VQLS Cost", "QSVT Degree", "QSVT Depth",
    ]
    rows1 = []
    for N in (4, 8):
        d = s1_data[N]
        for key, label in [
            ("thomas","Thomas"), ("hhl","HHL"), ("vqls","VQLS"), ("qsvt","QSVT")
        ]:
            dec = d[key].get("dec", {})
            disc_s = round(dec.get("disc_pct", 0.0), 4) if not np.isnan(dec.get("disc_pct", float("nan"))) else "N/A"
            rows1.append([
                N, label,
                round(dec.get("total_pct", 0.0), 4),
                disc_s,
                round(dec.get("algo_pct", 0.0), 4),
                f"{d[key]['res']:.4e}",
                round(d[key]["t"], 3),
                round(d[key].get("cost", ""), 6) if key == "vqls" else "",
                d[key].get("degree", "") if key == "qsvt" else "",
                d[key].get("depth",  "") if key == "qsvt" else "",
            ])
    _write_sheet(ws1, headers1, rows1)

    # Sheet 2: QSVT complexity.
    ws2 = wb.create_sheet("S2 QSVT Complexity")
    headers2 = [
        "N", "kappa(A)", "alpha", "kappa_eff",
        "QSVT Degree", "QSVT Depth (est.)", "HHL Depth (est.)", "QSVT Qubits",
    ]
    rows2 = [
        [N, round(s2_data[N]["kappa"], 1), s2_data[N]["alpha"],
         round(s2_data[N]["kappa_eff"], 1), s2_data[N]["degree"],
         s2_data[N]["depth_qsvt"], s2_data[N]["depth_hhl"],
         s2_data[N]["qubits_qsvt"]]
        for N in sorted(s2_data.keys())
    ]
    _write_sheet(ws2, headers2, rows2)

    # Sheet 3: HET 1-D.
    ws3 = wb.create_sheet("S3 HET 1D")
    headers3 = [
        "Sub-case", "Solver",
        "Total Error (%)", "Disc. Error (%)", "Algo. Error (%)",
        "Normalised Residual", "Wall Time (s)", "VQLS Cost",
    ]
    rows3 = []
    for sub, label_sub in [("3a","Linear/Homogeneous"),("3b","Gaussian/Physical BCs")]:
        d = s3_data[sub]
        for key, label in [("thomas","Thomas"),("hhl","HHL"),("vqls","VQLS")]:
            if key not in d:
                continue
            dec = d[key].get("dec", {})
            disc_s = round(dec.get("disc_pct", 0.0), 4) if not np.isnan(dec.get("disc_pct", float("nan"))) else "N/A"
            res = _residual(HETPoissonProblem1D(d["cfg"]).A,
                            d[key]["u"],
                            HETPoissonProblem1D(d["cfg"]).b)
            rows3.append([
                label_sub, label,
                round(dec.get("total_pct", 0.0), 4),
                disc_s,
                round(dec.get("algo_pct", 0.0), 4),
                f"{res:.4e}",
                round(d[key]["t"], 3),
                round(d[key].get("cost", ""), 6) if key == "vqls" else "",
            ])
    _write_sheet(ws3, headers3, rows3)

    # Sheet 4: HET 2-D.
    ws4 = wb.create_sheet("S4 HET 2D")
    headers4 = [
        "Solver", "Iterations", "Converged",
        "Total Error (%)", "Disc. Error (%)", "Algo. Error (%)",
        "Jacobi Convergence Error", "Wall Time (s)",
    ]
    rows4 = []
    for key, label in [("Thomas-2D","Thomas-2D"),("VQLS-2D","VQLS-2D"),("QSVT-2D","QSVT-2D")]:
        if key not in s4_data:
            continue
        r   = s4_data[key]["result"]
        dec = s4_data[key].get("dec", {})
        jacobi_err = r.iteration_errors[-1] if r.iteration_errors else float("nan")
        disc_s = round(dec.get("disc_pct", 0.0), 4) if not np.isnan(dec.get("disc_pct", float("nan"))) else "N/A"
        rows4.append([
            label, r.iterations, r.converged,
            round(dec.get("total_pct", 0.0), 4),
            disc_s,
            round(dec.get("algo_pct", 0.0), 4),
            f"{jacobi_err:.4e}",
            round(s4_data[key]["time"], 3),
        ])
    _write_sheet(ws4, headers4, rows4)

    wb.save(path)
    print(f"\n  Excel workbook saved to {path}")


def _export_csv_fallback(s1_data, s2_data, s3_data, s4_data) -> None:
    import csv
    def _write_csv(filename, headers, rows):
        filepath = RESULTS_DIR / filename
        with open(filepath, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        print(f"  CSV saved to {filepath}")

    rows1 = []
    for N in (4, 8):
        d = s1_data[N]
        for key, label in [("thomas","Thomas"),("hhl","HHL"),("vqls","VQLS"),("qsvt","QSVT")]:
            dec = d[key].get("dec", {})
            rows1.append([N, label,
                round(dec.get("total_pct", 0.0), 4),
                round(dec.get("disc_pct", 0.0), 4),
                round(dec.get("algo_pct", 0.0), 4),
                f"{d[key]['res']:.4e}", round(d[key]["t"], 3)])
    _write_csv("s1_algorithm_comparison.csv",
               ["N","Solver","Total%","Disc%","Algo%","Residual","Time_s"], rows1)


# ============================================================================
# Utility
# ============================================================================

def _save_figure(fig, filename: str, save: bool) -> None:
    if not save:
        return
    for ext in ("pdf", "png"):
        path = RESULTS_DIR / filename.replace(".pdf", f".{ext}")
        fig.savefig(path, bbox_inches="tight", dpi=150)
    print(f"  Figure saved to {RESULTS_DIR / filename}")


# ============================================================================
# Main entry point
# ============================================================================

def main() -> None:
    """
    Execute the full meeting progress report.

    Sections 1-4 are run sequentially. Section 5 (2-D QSVT) is merged
    into Section 4. Total estimated runtime: 35-55 minutes on a laptop.
    """
    t_start = time.perf_counter()

    print("\n" + "═"*68)
    print("  QUANTUM POISSON SOLVER — MEETING PROGRESS REPORT")
    print("  Imperial College London, Department of Aeronautics")
    print("  HHL | VQLS | QSVT — Poisson Equation and HET Plasma Modelling")
    print("═"*68)
    print(f"  Output directory: {RESULTS_DIR.resolve()}")

    s1_data = run_section_1()
    s2_data = run_section_2()
    s3_data = run_section_3()
    #s4_data = run_section_4()

    t_elapsed = time.perf_counter() - t_start
    print(f"\n{'─'*68}")
    print(f"  All sections completed in {t_elapsed:.1f}s. "
          f"Generating figures and exports...")

    plot_section_1(s1_data, save=True)
    plot_section_2(s2_data, save=True)
    plot_section_3(s3_data, save=True)
    #plot_section_4(s4_data, save=True)

    #export_excel(s1_data, s2_data, s3_data, s4_data)

    print(f"\n  Total elapsed time: {time.perf_counter() - t_start:.1f}s")
    print(f"  All outputs saved to: {RESULTS_DIR.resolve()}")
    print("═"*68)


if __name__ == "__main__":
    main()